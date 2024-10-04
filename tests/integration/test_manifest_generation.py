"""
This module is responsible for running through the "Manifest Generation" portion of
the schematic API test plan found here: <https://sagebionetworks.jira.com/wiki/spaces/SCHEM/pages/3055779846/Schematic+API+test+plan>.

There are a small number of items that need to be manually verified, and these are
noted in the test function docstrings.
"""
import os
from io import BytesIO

import pytest
import requests
from flask.testing import FlaskClient
from openpyxl import load_workbook

LIGHT_BLUE = "FFEAF7F9"  # Required cell
GRAY = "FFE0E0E0"  # Header cell
WHITE = "00000000"  # Optional cell

manual_test_verification_enabled = (
    os.environ.get("MANUAL_TEST_VERIFICATION", "false").lower() == "true"
)
use_deployed_schematic_api_server = (
    os.environ.get("USE_DEPLOYED_SCHEMATIC_API_SERVER", "false").lower() == "true"
)
schematic_api_server_url = os.environ.get(
    "SCHEMATIC_API_SERVER_URL", "http://localhost:3001"
)
local_flask_instance = not use_deployed_schematic_api_server or (
    use_deployed_schematic_api_server and "localhost" in schematic_api_server_url
)


class TestManifestGeneration:
    @pytest.mark.manual_verification_required
    def test_single_manifest_generation_excel(
        self,
        manual_test_verification_path: str,
        flask_client: FlaskClient,
        syn_token: str,
    ) -> None:
        """
        Download a manifest from the Schematic API and verify that it is a valid Excel
        file. We are validating the following:

        - The first row of the Excel file contains the column headers
        - The first row is locked on scroll
        - Each cell in the first row has a comment "TBD"
        - The cell corresponding to "Sex" in Sheet1 has a dropdown list with values from Sheet2!B2:B4
        - The cell corresponding to "Diagnosis" in Sheet1 has a dropdown list with values from Sheet2!D2:D3
        - The cell corresponding to "Cancer Type" in Sheet1 has a dropdown list with values from Sheet2!F2:F6
        - The workbook contains two sheets: "Sheet1" and "Sheet2"
        - "Sheet2" is hidden
        - The values in "Sheet2" are as expected
        - The fill colors of the first row cells are as expected
        - Conditional formatting is functioning as expected


        Manual verification steps:

        - Open the Excel file prefixed with TestManifestGeneration_test_single_manifest_generation_excel.xlsx
        - When Diagnosis = "Cancer", [Cancer Type, Family History] is Light Blue (Required)
        """
        # GIVEN a valid example manifest to generate
        url = f"{schematic_api_server_url}/v1/manifest/generate"
        params = {
            "schema_url": "https://raw.githubusercontent.com/Sage-Bionetworks/schematic/develop/tests/data/example.model.jsonld",
            "title": "Example",
            "data_type": "Patient",
            "use_annotations": "false",
            "output_format": "excel",
            "strict_validation": "true",
            "data_model_labels": "class_label",
        }
        headers = {"accept": "application/json", "Authorization": f"Bearer {syn_token}"}

        # WHEN we make a request to the Schematic API
        response = (
            requests.get(url, headers=headers, params=params, timeout=300)
            if use_deployed_schematic_api_server
            else flask_client.get(url, query_string=params, headers=headers)
        )

        # THEN we expect a successful response
        assert response.status_code == 200

        # Load the response content into memory
        content = BytesIO(
            response.content if use_deployed_schematic_api_server else response.data
        )
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert sheet1[f"{columns['Patient ID']}1"].value == "Patient ID"
        assert sheet1[f"{columns['Patient ID']}2"].value is None

        assert sheet1[f"{columns['Sex']}1"].value == "Sex"
        assert sheet1[f"{columns['Sex']}2"].value is None

        assert sheet1[f"{columns['Year of Birth']}1"].value == "Year of Birth"
        assert sheet1[f"{columns['Year of Birth']}2"].value is None

        assert sheet1[f"{columns['Diagnosis']}1"].value == "Diagnosis"
        assert sheet1[f"{columns['Diagnosis']}2"].value is None

        assert sheet1[f"{columns['Component']}1"].value == "Component"
        assert sheet1[f"{columns['Component']}2"].value == "Patient"

        assert sheet1[f"{columns['Cancer Type']}1"].value == "Cancer Type"
        assert sheet1[f"{columns['Cancer Type']}2"].value is None

        assert sheet1[f"{columns['Family History']}1"].value == "Family History"
        assert sheet1[f"{columns['Family History']}2"].value is None

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['Family History']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Component",
            "Cancer Type",
            "Family History",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND the comment in "Family History" cell is as expected
        assert (
            sheet1[f"{columns['Family History']}2"].comment.text
            == "Please enter applicable comma-separated items selected from the set of allowable terms for this attribute. See our data standards for allowable terms"
        )

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        sex_validation = None
        diagnosis_validation = None
        cancer_type_validation = None
        for dv in data_validations:
            if f"{columns['Sex']}2" in dv.sqref:
                sex_validation = dv
                continue
            elif f"{columns['Diagnosis']}2" in dv.sqref:
                diagnosis_validation = dv
                continue
            elif f"{columns['Cancer Type']}2" in dv.sqref:
                cancer_type_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert sex_validation is not None
        assert sex_validation.type == "list"
        assert sex_validation.formula1 == "Sheet2!$B$2:$B$4"

        assert diagnosis_validation is not None
        assert diagnosis_validation.type == "list"
        assert diagnosis_validation.formula1 == "Sheet2!$D$2:$D$3"

        assert cancer_type_validation is not None
        assert cancer_type_validation.type == "list"
        assert cancer_type_validation.formula1 == "Sheet2!$F$2:$F$6"

        # AND the fill colors are as expected
        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == WHITE

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['Diagnosis']}2"].value = "Cancer"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert workbook.sheetnames == ["Sheet1", "Sheet2"]

        sheet2 = workbook["Sheet2"]

        # AND the second sheet is hidden
        assert sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert sheet2["A1"].value == "Patient ID"
        assert sheet2["A2"].value is None
        assert sheet2["A3"].value is None
        assert sheet2["A4"].value is None
        assert sheet2["A5"].value is None
        assert sheet2["A6"].value is None

        assert sheet2["B1"].value == "Sex"
        assert sheet2["B2"].value == "Female"
        assert sheet2["B3"].value == "Male"
        assert sheet2["B4"].value == "Other"
        assert sheet2["B5"].value is None
        assert sheet2["B6"].value is None

        assert sheet2["C1"].value == "Year of Birth"
        assert sheet2["C2"].value is None
        assert sheet2["C3"].value is None
        assert sheet2["C4"].value is None
        assert sheet2["C5"].value is None
        assert sheet2["C6"].value is None

        assert sheet2["D1"].value == "Diagnosis"
        assert sheet2["D2"].value == "Cancer"
        assert sheet2["D3"].value == "Healthy"
        assert sheet2["D4"].value is None
        assert sheet2["D5"].value is None
        assert sheet2["D6"].value is None

        assert sheet2["E1"].value == "Component"
        assert sheet2["E2"].value is None
        assert sheet2["E3"].value is None
        assert sheet2["E4"].value is None
        assert sheet2["E5"].value is None
        assert sheet2["E6"].value is None

        assert sheet2["F1"].value == "Cancer Type"
        assert sheet2["F2"].value == "Breast"
        assert sheet2["F3"].value == "Colorectal"
        assert sheet2["F4"].value == "Lung"
        assert sheet2["F5"].value == "Prostate"
        assert sheet2["F6"].value == "Skin"

        assert sheet2["G1"].value == "Family History"
        assert sheet2["G2"].value == "Breast"
        assert sheet2["G3"].value == "Colorectal"
        assert sheet2["G4"].value == "Lung"
        assert sheet2["G5"].value == "Prostate"
        assert sheet2["G6"].value == "Skin"

        # AND there are no more columns in the second sheet
        assert sheet2["H1"].value is None

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    manual_test_verification_path,
                    "TestManifestGeneration_test_single_manifest_generation_excel.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_single_manifest_generation_google_sheet_with_annotations(
        self,
        manual_test_verification_path: str,
        flask_client: FlaskClient,
        syn_token: str,
    ) -> None:
        """
        Download a manifest with annotations from the Schematic API and verify that it is a valid Google
        Sheet. We are validating the following:

        - The first row of the Google Sheet contains the column headers
        - The first row is locked on scroll
        - Each cell A-F in the first row has a comment "TBD"
        - Each cell G-M in the first row does not have a comment
        - The cell corresponding to "File Format" in Sheet1 has a dropdown list with values from Sheet2!C2:C5
        - The cell corresponding to "Genome Build" in Sheet1 has a dropdown list with values from Sheet2!E2:E5
        - The fill colors of the first row cells are as expected
        - The workbook contains two sheets: "Sheet1" and "Sheet2"
        - "Sheet2" is hidden
        - The values in "Sheet1" are as expected
        - The values in "Sheet2" are as expected

        Manual verification steps:
        - When File Format = "BAM", [Genome Build] is Light Blue (Required)
        - When File Format = "CRAM", [Genome Build, Genome FASTA] is Light Blue (Required)
        - When File Format = "FASTQ", [Genome Build] is White (Optional)
        """
        # GIVEN a valid example manifest to generate
        url = f"{schematic_api_server_url}/v1/manifest/generate"
        params = {
            "schema_url": "https://raw.githubusercontent.com/Sage-Bionetworks/schematic/develop/tests/data/example.model.jsonld",
            "title": "Example",
            "data_type": "BulkRNA-seqAssay",
            "use_annotations": "true",
            "dataset_id": "syn63561056",
            "asset_view": "syn63561086",
            "output_format": "google_sheet",
            "strict_validation": "true",
            "data_model_labels": "class_label",
        }
        headers = {"accept": "application/json", "Authorization": f"Bearer {syn_token}"}

        # WHEN we make a request to the Schematic API
        response = (
            requests.get(url, headers=headers, params=params, timeout=300)
            if use_deployed_schematic_api_server
            else flask_client.get(url, query_string=params, headers=headers)
        )

        # THEN we expect a successful response
        assert response.status_code == 200

        # Load the Google Sheets URL from the response
        response_content = (
            response.json() if use_deployed_schematic_api_server else response.json
        )
        assert len(response_content) == 1
        google_sheet_url = response_content[0]
        assert (
            google_sheet_url is not None
        ), "No Google Sheets URL found in the response"

        # Download the Google Sheets content as an Excel file and load into openpyxl
        export_url = f"{google_sheet_url}/export?format=xlsx"
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert columns["Filename"] is not None
        assert columns["Sample ID"] is not None
        assert columns["File Format"] is not None
        assert columns["Component"] is not None
        assert columns["Genome Build"] is not None
        assert columns["Genome FASTA"] is not None
        assert columns["eTag"] is not None
        assert columns["key_bool"] is not None
        assert columns["key_int"] is not None
        assert columns["key_float"] is not None
        assert columns["key_str"] is not None
        assert columns["key_datetime"] is not None
        assert columns["entityId"] is not None

        if local_flask_instance:
            assert (
                sheet1[f"{columns['Filename']}2"].value
                == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 1/sample A.txt"
            )
        else:
            assert (
                sheet1[f"{columns['Filename']}2"].value
                == "test dataset/test dataset 1/sample A.txt"
            )

        assert sheet1[f"{columns['Sample ID']}2"].value is None
        assert sheet1[f"{columns['File Format']}2"].value is None
        assert sheet1[f"{columns['Component']}2"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}2"].value is None
        assert sheet1[f"{columns['Genome FASTA']}2"].value is None
        assert sheet1[f"{columns['eTag']}2"].value is not None  # eTag
        assert isinstance(sheet1[f"{columns['key_bool']}2"].value, bool)
        assert sheet1[f"{columns['key_bool']}2"].value
        assert sheet1[f"{columns['key_int']}2"].value == 6
        assert sheet1[f"{columns['key_float']}2"].value == 80
        assert sheet1[f"{columns['key_str']}2"].value == "New Value"
        assert sheet1[f"{columns['key_datetime']}2"].value is not None  # key_datetime
        assert sheet1[f"{columns['entityId']}2"].value == "syn63561081"

        if local_flask_instance:
            assert (
                sheet1[f"{columns['Filename']}3"].value
                == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 2/sample B.txt"
            )
        else:
            assert (
                sheet1[f"{columns['Filename']}3"].value
                == "test dataset/test dataset 2/sample B.txt"
            )
        assert sheet1[f"{columns['Sample ID']}3"].value is None
        assert sheet1[f"{columns['File Format']}3"].value is None
        assert sheet1[f"{columns['Component']}3"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}3"].value is None
        assert sheet1[f"{columns['Genome FASTA']}3"].value is None
        assert sheet1[f"{columns['eTag']}3"].value is not None  # eTag
        assert sheet1[f"{columns['key_bool']}3"].value is None
        assert sheet1[f"{columns['key_int']}3"].value is None
        assert sheet1[f"{columns['key_float']}3"].value is None
        assert sheet1[f"{columns['key_str']}3"].value is None
        assert sheet1[f"{columns['key_datetime']}3"].value is None
        assert sheet1[f"{columns['entityId']}3"].value == "syn63561082"

        if local_flask_instance:
            assert (
                sheet1[f"{columns['Filename']}4"].value
                == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 3/sample C.txt"
            )
        else:
            assert (
                sheet1[f"{columns['Filename']}4"].value
                == "test dataset/test dataset 3/sample C.txt"
            )

        assert sheet1[f"{columns['Sample ID']}4"].value is None
        assert sheet1[f"{columns['File Format']}4"].value is None
        assert sheet1[f"{columns['Component']}4"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}4"].value is None
        assert sheet1[f"{columns['Genome FASTA']}4"].value is None
        assert sheet1[f"{columns['eTag']}4"].value is not None  # eTag
        assert sheet1[f"{columns['key_bool']}4"].value is None
        assert sheet1[f"{columns['key_int']}4"].value is None
        assert sheet1[f"{columns['key_float']}4"].value is None
        assert sheet1[f"{columns['key_str']}4"].value is None
        assert sheet1[f"{columns['key_datetime']}4"].value is None
        assert sheet1[f"{columns['entityId']}4"].value == "syn63561085"

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['entityId']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each of these cells in the first row has a comment "TBD"
        for col in [
            "Filename",
            "Sample ID",
            "File Format",
            "Component",
            "Genome Build",
            "Genome FASTA",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND each of these cells in the first row do not have a comment
        for col in [
            "eTag",
            "key_bool",
            "key_int",
            "key_float",
            "key_str",
            "key_datetime",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].comment is None

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        file_format_validation = None
        genome_build_validation = None
        for dv in data_validations:
            if f"{columns['File Format']}2" in dv.sqref:
                file_format_validation = dv
                continue
            elif f"{columns['Genome Build']}2" in dv.sqref:
                genome_build_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert file_format_validation is not None
        assert file_format_validation.type == "list"
        assert (
            file_format_validation.formula1
            == f"Sheet2!${columns['File Format']}$2:${columns['File Format']}$5"
        )

        assert genome_build_validation is not None
        assert genome_build_validation.type == "list"
        assert (
            genome_build_validation.formula1
            == f"Sheet2!${columns['Genome Build']}$2:${columns['Genome Build']}$5"
        )

        # AND the fill colors are as expected
        for col in ["Filename", "Sample ID", "File Format", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in [
            "Genome Build",
            "Genome FASTA",
            "eTag",
            "key_bool",
            "key_int",
            "key_float",
            "key_str",
            "key_datetime",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['File Format']}2"].value = "BAM"
        workbook["Sheet1"][f"{columns['File Format']}3"].value = "CRAM"
        workbook["Sheet1"][f"{columns['File Format']}4"].value = "FASTQ"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert sheet2["A1"].value == "Filename"
        assert sheet2["B1"].value == "Sample ID"
        assert sheet2["C1"].value == "File Format"
        assert sheet2["D1"].value == "Component"
        assert sheet2["E1"].value == "Genome Build"
        assert sheet2["F1"].value == "Genome FASTA"

        assert sheet2["A2"].value is None
        assert sheet2["B2"].value is None
        assert sheet2["C2"].value == "BAM"
        assert sheet2["D2"].value is None
        assert sheet2["E2"].value == "GRCh37"
        assert sheet2["F2"].value is None

        assert sheet2["A3"].value is None
        assert sheet2["B3"].value is None
        assert sheet2["C3"].value == "CRAM"
        assert sheet2["D3"].value is None
        assert sheet2["E3"].value == "GRCh38"
        assert sheet2["F3"].value is None

        assert sheet2["A4"].value is None
        assert sheet2["B4"].value is None
        assert sheet2["C4"].value == "CSV/TSV"
        assert sheet2["D4"].value is None
        assert sheet2["E4"].value == "GRCm38"
        assert sheet2["F4"].value is None

        assert sheet2["A5"].value is None
        assert sheet2["B5"].value is None
        assert sheet2["C5"].value == "FASTQ"
        assert sheet2["D5"].value is None
        assert sheet2["E5"].value == "GRCm39"
        assert sheet2["F5"].value is None

        # AND there are no more columns in the second sheet
        assert sheet2["G1"].value is None

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    manual_test_verification_path,
                    "TestManifestGeneration_test_single_manifest_generation_google_sheet_with_annotations.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_single_manifest_generation_google_sheet_no_annotations(
        self,
        manual_test_verification_path: str,
        flask_client: FlaskClient,
        syn_token: str,
    ) -> None:
        """
        Download a manifest without annotations from the Schematic API and verify that it is a valid Google
        Sheet. We are validating the following:

        - The first row of the Google Sheet contains the column headers
        - The first row is locked on scroll
        - Each cell A-F in the first row has a comment "TBD"
        - Cell G in the first row does not have a comment
        - The second cell in the "File Format" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The second cell in the "Genome Build" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The fill colors of the first row cells are as expected
        - The workbook contains two sheets: "Sheet1" and "Sheet2"
        - "Sheet2" is hidden
        - The values in "Sheet1" are as expected
        - The values in "Sheet2" are as expected

        Manual verification steps:
        - When File Format = "BAM", [Genome Build] is Light Blue (Required)
        - When File Format = "CRAM", [Genome Build, Genome FASTA] is Light Blue (Required)
        - When File Format = "FASTQ", [Genome Build] is White (Optional)
        """
        url = f"{schematic_api_server_url}/v1/manifest/generate"
        # GIVEN a valid request to the Schematic API to generate a Google Sheet manifest without annotations
        params = {
            "schema_url": "https://raw.githubusercontent.com/Sage-Bionetworks/schematic/develop/tests/data/example.model.jsonld",
            "title": "Example",
            "data_type": "BulkRNA-seqAssay",
            "use_annotations": "false",
            "dataset_id": "syn63561056",
            "asset_view": "syn63561086",
            "output_format": "google_sheet",
            "strict_validation": "true",
            "data_model_labels": "class_label",
        }
        headers = {"accept": "application/json", "Authorization": f"Bearer {syn_token}"}
        # WHEN we make a request to the Schematic API
        response = (
            requests.get(url, headers=headers, params=params, timeout=300)
            if use_deployed_schematic_api_server
            else flask_client.get(url, query_string=params, headers=headers)
        )

        # THEN we expect a successful response
        assert response.status_code == 200

        # Load the Google Sheets URL from the response
        response_content = (
            response.json() if use_deployed_schematic_api_server else response.json
        )
        assert len(response_content) == 1, "Expected a single URL in the response"
        google_sheet_url = response_content[0]
        assert (
            google_sheet_url is not None
        ), "No Google Sheets URL found in the response"

        # Convert the Google Sheets URL to an export URL for Excel format
        export_url = f"{google_sheet_url}/export?format=xlsx"

        # AND we should be able to download the manifest as an Excel file
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of sheet1 is as expected
        assert columns["Filename"] is not None
        assert columns["Sample ID"] is not None
        assert columns["File Format"] is not None
        assert columns["Component"] is not None
        assert columns["Genome Build"] is not None
        assert columns["Genome FASTA"] is not None
        assert columns["entityId"] is not None

        assert (
            sheet1[f"{columns['Filename']}2"].value
            == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 1/sample A.txt"
        )
        assert sheet1[f"{columns['Sample ID']}2"].value is None
        assert sheet1[f"{columns['File Format']}2"].value is None
        assert sheet1[f"{columns['Component']}2"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}2"].value is None
        assert sheet1[f"{columns['Genome FASTA']}2"].value is None
        assert sheet1[f"{columns['entityId']}2"].value == "syn63561081"

        assert (
            sheet1[f"{columns['Filename']}3"].value
            == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 2/sample B.txt"
        )
        assert sheet1[f"{columns['Sample ID']}3"].value is None
        assert sheet1[f"{columns['File Format']}3"].value is None
        assert sheet1[f"{columns['Component']}3"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}3"].value is None
        assert sheet1[f"{columns['Genome FASTA']}3"].value is None
        assert sheet1[f"{columns['entityId']}3"].value == "syn63561082"

        assert (
            sheet1[f"{columns['Filename']}4"].value
            == "Manifest generation - Manual test - generate an existing manifest/test dataset/test dataset 3/sample C.txt"
        )
        assert sheet1[f"{columns['Sample ID']}4"].value is None
        assert sheet1[f"{columns['File Format']}4"].value is None
        assert sheet1[f"{columns['Component']}4"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}4"].value is None
        assert sheet1[f"{columns['Genome FASTA']}4"].value is None
        assert sheet1[f"{columns['entityId']}4"].value == "syn63561085"

        # AND there are no more columns in the sheet
        assert sheet1[f"{columns['entityId']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each of these cells in the first row has a comment "TBD"
        for col in [
            "Filename",
            "Sample ID",
            "File Format",
            "Component",
            "Genome Build",
            "Genome FASTA",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND the entityId column in the first row does not have a comment
        assert sheet1[f"{columns['entityId']}1"].comment is None

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        file_format_validation = None
        genome_build_validation = None
        for dv in data_validations:
            if f"{columns['File Format']}2" in dv.sqref:
                file_format_validation = dv
                continue
            elif f"{columns['Genome Build']}2" in dv.sqref:
                genome_build_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert file_format_validation is not None
        assert file_format_validation.type == "list"
        assert (
            file_format_validation.formula1
            == f"Sheet2!${columns['File Format']}$2:${columns['File Format']}$5"
        )

        assert genome_build_validation is not None
        assert genome_build_validation.type == "list"
        assert (
            genome_build_validation.formula1
            == f"Sheet2!${columns['Genome Build']}$2:${columns['Genome Build']}$5"
        )

        # AND the fill colors are as expected
        for col in ["Filename", "Sample ID", "File Format", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in [
            "Genome Build",
            "Genome FASTA",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['File Format']}2"].value = "BAM"
        workbook["Sheet1"][f"{columns['File Format']}3"].value = "CRAM"
        workbook["Sheet1"][f"{columns['File Format']}4"].value = "FASTQ"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert sheet2["A1"].value == "Filename"
        assert sheet2["B1"].value == "Sample ID"
        assert sheet2["C1"].value == "File Format"
        assert sheet2["D1"].value == "Component"
        assert sheet2["E1"].value == "Genome Build"
        assert sheet2["F1"].value == "Genome FASTA"

        assert sheet2["A2"].value is None
        assert sheet2["B2"].value is None
        assert sheet2["C2"].value == "BAM"
        assert sheet2["D2"].value is None
        assert sheet2["E2"].value == "GRCh37"
        assert sheet2["F2"].value is None

        assert sheet2["A3"].value is None
        assert sheet2["B3"].value is None
        assert sheet2["C3"].value == "CRAM"
        assert sheet2["D3"].value is None
        assert sheet2["E3"].value == "GRCh38"
        assert sheet2["F3"].value is None

        assert sheet2["A4"].value is None
        assert sheet2["B4"].value is None
        assert sheet2["C4"].value == "CSV/TSV"
        assert sheet2["D4"].value is None
        assert sheet2["E4"].value == "GRCm38"
        assert sheet2["F4"].value is None

        assert sheet2["A5"].value is None
        assert sheet2["B5"].value is None
        assert sheet2["C5"].value == "FASTQ"
        assert sheet2["D5"].value is None
        assert sheet2["E5"].value == "GRCm39"
        assert sheet2["F5"].value is None

        # AND there are no more columns in the second sheet
        assert sheet2["G1"].value is None

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    manual_test_verification_path,
                    "TestManifestGeneration_test_single_manifest_generation_google_sheet_no_annotations.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_manifest_generation_multiple_blank_google_sheets(
        self,
        manual_test_verification_path: str,
        flask_client: FlaskClient,
        syn_token: str,
    ) -> None:
        """
        Download two blank manifests from the Schematic API and verify that they are valid Google Sheets.
        We are validating the following:

        For the Patient Google Sheet:
        - The first row of the Google Sheet contains the column headers
        - The first row is locked on scroll
        - Each cell A-G in the first row has a comment "TBD"
        - Cell H in the first row does not have a comment
        - The "Sex" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The "Diagnosis" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The "Cancer Type" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The "Family History" column in "Sheet1" has a comment that starts with "Please enter applicable comma-separated items"
        - The fill colors of the first row cells are as expected
        - The workbook contains two sheets: "Sheet1" and "Sheet2"
        - The second sheet is hidden
        - The values in "Sheet1" are as expected
        - The values in "Sheet2" are as expected
        Manual verification steps:
        - When Diagnosis = "Cancer", [Cancer Type, Family History] are Light Blue (Required)

        For the Bulk RNA-seq Assay Google Sheet:
        - The first row of the Google Sheet contains the column headers
        - The first row is locked on scroll
        - Each cell A-F in the first row has a comment "TBD"
        - Each cell G-M in the first row does not have a comment
        - The "File Format" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The "Genome Build" column in "Sheet1" has a dropdown list with the correct values from "Sheet2"
        - The fill colors of the first row cells are as expected
        - The workbook contains two sheets: "Sheet1" and "Sheet2"
        - "Sheet2" is hidden
        - The values in "Sheet1" are as expected
        - The values in "Sheet2" are as expected

        Manual verification steps:
        - When File Format = "BAM", [Genome Build] is Light Blue (Required)
        - When File Format = "CRAM", [Genome Build, Genome FASTA] is Light Blue (Required)
        - When File Format = "FASTQ", [Genome Build] is White (Optional)
        """
        url = f"{schematic_api_server_url}/v1/manifest/generate"
        # GIVEN a valid request to the Schematic API to generate two blank Google Sheets manifests
        params = {
            "schema_url": "https://raw.githubusercontent.com/Sage-Bionetworks/schematic/develop/tests/data/example.model.jsonld",
            "title": "Example",
            "data_type": "Patient,BulkRNA-seqAssay",
            "use_annotations": "false",
            "output_format": "google_sheet",
            "strict_validation": "true",
            "data_model_labels": "class_label",
        }
        headers = {"accept": "application/json", "Authorization": f"Bearer {syn_token}"}
        # WHEN we make a request to the Schematic API
        response = (
            requests.get(url, headers=headers, params=params, timeout=300)
            if use_deployed_schematic_api_server
            else flask_client.get(url, query_string=params, headers=headers)
        )

        # THEN we expect a successful response
        assert response.status_code == 200

        # Load the Google Sheets URLs from the response
        response_content = (
            response.json() if use_deployed_schematic_api_server else response.json
        )
        assert (
            len(response_content) == 2
        ), "Expected two Google Sheets URLs in the response"
        google_sheet_urls = response_content
        assert (
            google_sheet_urls is not None
        ), "No Google Sheets URLs found in the response"

        # Convert the Google Sheets URLs to export URLs for Excel format
        export_urls = [f"{url}/export?format=xlsx" for url in google_sheet_urls]
        patient_export_url = export_urls[0]
        rna_seq_export_url = export_urls[1]

        # AND we should be able to download the patient manifest as an Excel file
        patient_response = requests.get(patient_export_url)
        assert patient_response.status_code == 200
        patient_content = BytesIO(patient_response.content)
        patient_workbook = load_workbook(patient_content)
        patient_sheet1 = patient_workbook["Sheet1"]
        patient_sheet2 = patient_workbook["Sheet2"]

        # Track column positions
        patient_columns = {cell.value: cell.column_letter for cell in patient_sheet1[1]}

        # AND the content of sheet1 is as expected
        assert patient_sheet1[f"{patient_columns['Patient ID']}1"].value == "Patient ID"
        assert patient_sheet1[f"{patient_columns['Sex']}1"].value == "Sex"
        assert (
            patient_sheet1[f"{patient_columns['Year of Birth']}1"].value
            == "Year of Birth"
        )
        assert patient_sheet1[f"{patient_columns['Diagnosis']}1"].value == "Diagnosis"
        assert patient_sheet1[f"{patient_columns['Component']}1"].value == "Component"
        assert (
            patient_sheet1[f"{patient_columns['Cancer Type']}1"].value == "Cancer Type"
        )
        assert (
            patient_sheet1[f"{patient_columns['Family History']}1"].value
            == "Family History"
        )

        assert patient_sheet1[f"{patient_columns['Patient ID']}2"].value is None
        assert patient_sheet1[f"{patient_columns['Sex']}2"].value is None
        assert patient_sheet1[f"{patient_columns['Year of Birth']}2"].value is None
        assert patient_sheet1[f"{patient_columns['Diagnosis']}2"].value is None
        assert patient_sheet1[f"{patient_columns['Component']}2"].value == "Patient"
        assert patient_sheet1[f"{patient_columns['Cancer Type']}2"].value is None
        assert patient_sheet1[f"{patient_columns['Family History']}2"].value is None

        # AND there are no more columns in the first sheet
        assert (
            patient_sheet1[f"{patient_columns['Family History']}1"]
            .offset(column=1)
            .value
            is None
        )

        # AND the first row is locked on scroll
        assert patient_sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Component",
            "Cancer Type",
            "Family History",
        ]:
            assert patient_sheet1[f"{patient_columns[col]}1"].comment.text == "TBD"

        # AND the comment in "Family History" cell is as expected
        assert (
            patient_sheet1[f"{patient_columns['Family History']}2"].comment.text
            == "Please enter applicable comma-separated items selected from the set of allowable terms for this attribute. See our data standards for allowable terms"
        )

        # AND the dropdown lists exist and are as expected
        data_validations = patient_sheet1.data_validations.dataValidation
        sex_validation = None
        diagnosis_validation = None
        cancer_type_validation = None
        for dv in data_validations:
            if f"{patient_columns['Sex']}2" in dv.sqref:
                sex_validation = dv
                continue
            elif f"{patient_columns['Diagnosis']}2" in dv.sqref:
                diagnosis_validation = dv
                continue
            elif f"{patient_columns['Cancer Type']}2" in dv.sqref:
                cancer_type_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert sex_validation is not None
        assert sex_validation.type == "list"
        assert (
            sex_validation.formula1
            == f"Sheet2!${patient_columns['Sex']}$2:${patient_columns['Sex']}$4"
        )

        assert diagnosis_validation is not None
        assert diagnosis_validation.type == "list"
        assert (
            diagnosis_validation.formula1
            == f"Sheet2!${patient_columns['Diagnosis']}$2:${patient_columns['Diagnosis']}$3"
        )

        assert cancer_type_validation is not None
        assert cancer_type_validation.type == "list"
        assert (
            cancer_type_validation.formula1
            == f"Sheet2!${patient_columns['Cancer Type']}$2:${patient_columns['Cancer Type']}$6"
        )

        # AND the fill colors are as expected
        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert (
                patient_sheet1[f"{patient_columns[col]}1"].fill.start_color.index
                == LIGHT_BLUE
            )

        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert (
                patient_sheet1[f"{patient_columns[col]}2"].fill.start_color.index
                == LIGHT_BLUE
            )

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert (
                patient_sheet1[f"{patient_columns[col]}1"].fill.start_color.index
                == GRAY
            )

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert (
                patient_sheet1[f"{patient_columns[col]}2"].fill.start_color.index
                == WHITE
            )

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        patient_workbook["Sheet1"][f"{patient_columns['Diagnosis']}2"].value = "Cancer"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert patient_workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert patient_sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert patient_sheet2["A1"].value == "Patient ID"
        assert patient_sheet2["A2"].value is None
        assert patient_sheet2["A3"].value is None
        assert patient_sheet2["A4"].value is None
        assert patient_sheet2["A5"].value is None
        assert patient_sheet2["A6"].value is None

        assert patient_sheet2["B1"].value == "Sex"
        assert patient_sheet2["B2"].value == "Female"
        assert patient_sheet2["B3"].value == "Male"
        assert patient_sheet2["B4"].value == "Other"
        assert patient_sheet2["B5"].value is None
        assert patient_sheet2["B6"].value is None

        assert patient_sheet2["C1"].value == "Year of Birth"
        assert patient_sheet2["C2"].value is None
        assert patient_sheet2["C3"].value is None
        assert patient_sheet2["C4"].value is None
        assert patient_sheet2["C5"].value is None
        assert patient_sheet2["C6"].value is None

        assert patient_sheet2["D1"].value == "Diagnosis"
        assert patient_sheet2["D2"].value == "Cancer"
        assert patient_sheet2["D3"].value == "Healthy"
        assert patient_sheet2["D4"].value is None
        assert patient_sheet2["D5"].value is None
        assert patient_sheet2["D6"].value is None

        assert patient_sheet2["E1"].value == "Component"
        assert patient_sheet2["E2"].value is None
        assert patient_sheet2["E3"].value is None
        assert patient_sheet2["E4"].value is None
        assert patient_sheet2["E5"].value is None
        assert patient_sheet2["E6"].value is None

        assert patient_sheet2["F1"].value == "Cancer Type"
        assert patient_sheet2["F2"].value == "Breast"
        assert patient_sheet2["F3"].value == "Colorectal"
        assert patient_sheet2["F4"].value == "Lung"
        assert patient_sheet2["F5"].value == "Prostate"
        assert patient_sheet2["F6"].value == "Skin"

        assert patient_sheet2["G1"].value == "Family History"
        assert patient_sheet2["G2"].value == "Breast"
        assert patient_sheet2["G3"].value == "Colorectal"
        assert patient_sheet2["G4"].value == "Lung"
        assert patient_sheet2["G5"].value == "Prostate"
        assert patient_sheet2["G6"].value == "Skin"

        # AND there are no more columns in the second sheet
        assert patient_sheet2["H1"].value is None

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if manual_test_verification_enabled:
            patient_workbook.save(
                os.path.join(
                    manual_test_verification_path,
                    "TestManifestGeneration_test_multiple_blank_google_sheets_patient.xlsx",
                )
            )

        # AND we should be able to download the Bulk RNA-seq assay manifest as an Excel file
        rna_seq_response = requests.get(rna_seq_export_url)
        assert rna_seq_response.status_code == 200
        rna_seq_content = BytesIO(rna_seq_response.content)
        rna_seq_workbook = load_workbook(rna_seq_content)
        rna_seq_sheet1 = rna_seq_workbook["Sheet1"]
        rna_seq_sheet2 = rna_seq_workbook["Sheet2"]

        # Track column positions
        rna_seq_columns = {cell.value: cell.column_letter for cell in rna_seq_sheet1[1]}

        # AND the content of "Sheet1" is as expected
        assert rna_seq_columns["Filename"] is not None
        assert rna_seq_columns["Sample ID"] is not None
        assert rna_seq_columns["File Format"] is not None
        assert rna_seq_columns["Component"] is not None
        assert rna_seq_columns["Genome Build"] is not None
        assert rna_seq_columns["Genome FASTA"] is not None

        assert rna_seq_sheet1[f"{rna_seq_columns['Filename']}2"].value is None
        assert rna_seq_sheet1[f"{rna_seq_columns['Sample ID']}2"].value is None
        assert rna_seq_sheet1[f"{rna_seq_columns['File Format']}2"].value is None
        assert (
            rna_seq_sheet1[f"{rna_seq_columns['Component']}2"].value
            == "BulkRNA-seqAssay"
        )
        assert rna_seq_sheet1[f"{rna_seq_columns['Genome Build']}2"].value is None
        assert rna_seq_sheet1[f"{rna_seq_columns['Genome FASTA']}2"].value is None

        # AND there are no more columns in the sheet
        assert (
            rna_seq_sheet1[f"{rna_seq_columns['Genome FASTA']}1"].offset(column=1).value
            is None
        )

        # AND the first row is locked on scroll
        assert rna_seq_sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Filename",
            "Sample ID",
            "File Format",
            "Component",
            "Genome Build",
            "Genome FASTA",
        ]:
            assert rna_seq_sheet1[f"{rna_seq_columns[col]}1"].comment.text == "TBD"

        # AND the dropdown lists exist and are as expected
        data_validations = rna_seq_sheet1.data_validations.dataValidation
        file_format_validation = None
        genome_build_validation = None
        for dv in data_validations:
            if f"{rna_seq_columns['File Format']}2" in dv.sqref:
                file_format_validation = dv
                continue
            elif f"{rna_seq_columns['Genome Build']}2" in dv.sqref:
                genome_build_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert file_format_validation is not None
        assert file_format_validation.type == "list"
        assert (
            file_format_validation.formula1
            == f"Sheet2!${rna_seq_columns['File Format']}$2:${rna_seq_columns['File Format']}$5"
        )

        assert genome_build_validation is not None
        assert genome_build_validation.type == "list"
        assert (
            genome_build_validation.formula1
            == f"Sheet2!${rna_seq_columns['Genome Build']}$2:${rna_seq_columns['Genome Build']}$5"
        )

        # AND the fill colors are as expected
        for col in ["Filename", "Sample ID", "File Format", "Component"]:
            assert (
                rna_seq_sheet1[f"{rna_seq_columns[col]}1"].fill.start_color.index
                == LIGHT_BLUE
            )

        for col in [
            "Genome Build",
            "Genome FASTA",
        ]:
            assert (
                rna_seq_sheet1[f"{rna_seq_columns[col]}1"].fill.start_color.index
                == GRAY
            )
            assert (
                rna_seq_sheet1[f"{rna_seq_columns[col]}2"].fill.start_color.index
                == WHITE
            )

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        rna_seq_workbook["Sheet1"][f"{rna_seq_columns['File Format']}2"].value = "BAM"
        rna_seq_workbook["Sheet1"][f"{rna_seq_columns['File Format']}3"].value = "CRAM"
        rna_seq_workbook["Sheet1"][f"{rna_seq_columns['File Format']}4"].value = "FASTQ"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert rna_seq_workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert rna_seq_sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert rna_seq_sheet2["A1"].value == "Filename"
        assert rna_seq_sheet2["B1"].value == "Sample ID"
        assert rna_seq_sheet2["C1"].value == "File Format"
        assert rna_seq_sheet2["D1"].value == "Component"
        assert rna_seq_sheet2["E1"].value == "Genome Build"
        assert rna_seq_sheet2["F1"].value == "Genome FASTA"

        assert rna_seq_sheet2["A2"].value is None
        assert rna_seq_sheet2["B2"].value is None
        assert rna_seq_sheet2["C2"].value == "BAM"
        assert rna_seq_sheet2["D2"].value is None
        assert rna_seq_sheet2["E2"].value == "GRCh37"
        assert rna_seq_sheet2["F2"].value is None

        assert rna_seq_sheet2["A3"].value is None
        assert rna_seq_sheet2["B3"].value is None
        assert rna_seq_sheet2["C3"].value == "CRAM"
        assert rna_seq_sheet2["D3"].value is None
        assert rna_seq_sheet2["E3"].value == "GRCh38"
        assert rna_seq_sheet2["F3"].value is None

        assert rna_seq_sheet2["A4"].value is None
        assert rna_seq_sheet2["B4"].value is None
        assert rna_seq_sheet2["C4"].value == "CSV/TSV"
        assert rna_seq_sheet2["D4"].value is None
        assert rna_seq_sheet2["E4"].value == "GRCm38"
        assert rna_seq_sheet2["F4"].value is None

        assert rna_seq_sheet2["A5"].value is None
        assert rna_seq_sheet2["B5"].value is None
        assert rna_seq_sheet2["C5"].value == "FASTQ"
        assert rna_seq_sheet2["D5"].value is None
        assert rna_seq_sheet2["E5"].value == "GRCm39"
        assert rna_seq_sheet2["F5"].value is None

        # And there are no more columns in the second sheet
        assert rna_seq_sheet2["G1"].value is None

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if manual_test_verification_enabled:
            rna_seq_workbook.save(
                os.path.join(
                    manual_test_verification_path,
                    "TestManifestGeneration_test_multiple_blank_google_sheets_rna_seq.xlsx",
                )
            )
