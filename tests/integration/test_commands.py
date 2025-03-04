"""Tests for CLI commands"""

import os
import uuid
from io import BytesIO

import numpy as np
import pytest
import requests
from click.testing import CliRunner
from openpyxl import load_workbook

from schematic.configuration.configuration import CONFIG, Configuration
from schematic.manifest.commands import manifest
from schematic.models.commands import model
from schematic.utils.df_utils import read_csv
from tests.conftest import ConfigurationForTesting

LIGHT_BLUE = "FFEAF7F9"  # Required cell
GRAY = "FFE0E0E0"  # Header cell
WHITE = "00000000"  # Optional cell


@pytest.fixture(name="runner")
def fixture_runner() -> CliRunner:
    """Fixture for invoking command-line interfaces."""
    return CliRunner()


class TestSubmitCommand:
    """
    Tests for:
    - The command ran without error
    - The output contained a message indicating no validation errors
    - The output contained a message indicating the file was submitted
    """

    def test_submit_test_manifest(self, runner: CliRunner) -> None:
        """Tests for a successful submission"""
        try:
            result = runner.invoke(
                model,
                [
                    "--config",
                    "config_example.yml",
                    "submit",
                    "-mp",
                    "tests/data/mock_manifests/CLI_tests/CLI_biospecimen.csv",
                    "-vc",
                    "Biospecimen",
                    "-mrt",
                    "table_and_file",
                    "-d",
                    "syn23643250",
                    "--no-file_annotations_upload",
                ],
            )
            assert os.path.isfile("tests/data/example.Biospecimen.schema.json")

        finally:
            os.remove("tests/data/example.Biospecimen.schema.json")

        assert result.exit_code == 0
        assert "No validation errors occured during validation." in result.output
        assert (
            "File at 'tests/data/mock_manifests/CLI_tests/CLI_biospecimen.csv' was "
            "successfully associated with dataset 'syn23643250'."
        ) in result.output


class TestValidateCommand:
    """Tests the schematic/models/commands validate command"""

    def test_validate_valid_manifest(self, runner: CliRunner) -> None:
        """
        Tests for:
        - command has no (python) errors, has exit code 0
        - command output has success message
        - command output has no validation errors
        """
        result = runner.invoke(
            model,
            [
                "--config",
                "config_example.yml",
                "validate",
                "--manifest_path",
                "tests/data/mock_manifests/Valid_Test_Manifest.csv",
                "--data_type",
                "MockComponent",
                "--project_scope",
                "syn54126707",
            ],
        )
        # command has no (python) errors, has exit code 0
        assert result.exit_code == 0
        # command output has success message
        result_list = result.output.split("\n")
        assert (
            "Your manifest has been validated successfully. "
            "There are no errors in your manifest, "
            "and it can be submitted without any modifications."
        ) in result_list
        # command output has no validation errors
        errors = [errors for result in result_list if result.startswith("error")]

    def test_validate_invalid_manifest(self, runner: CliRunner) -> None:
        """
        Tests for:
        - command has no (python) errors, has exit code 0
        - command output includes error message: 'Random' is not a comma delimited string
        - command output includes error message: 'Random' is not one of
        """
        result = runner.invoke(
            model,
            [
                "--config",
                "config_example.yml",
                "validate",
                "--manifest_path",
                "tests/data/mock_manifests/CLI_tests/CLI_patient_invalid.csv",
                "--data_type",
                "Patient",
            ],
        )
        # command has no (python) errors, has exit code 0
        assert result.exit_code == 0
        # command output includes error message: 'Random' is not a comma delimited string
        assert result.output.split("\n")[3] == (
            "error: For attribute Family History in row 2 it does not appear "
            "as if you provided a comma delimited string. Please check your entry "
            "('Random'') and try again."
        )
        # command output includes error message: 'Random' is not one of
        # Note: the list of allowed values seems to have a random order so
        #   is not included in the test
        assert result.output.split("\n")[4].startswith("error: 'Random' is not one of")


class TestManifestCommand:
    """Tests the schematic/manifest/commands validate manifest command"""

    def test_generate_empty_csv_manifests(self, runner: CliRunner) -> None:
        """
        Tests for:
        - command has no errors, has exit code 0
        - command output has file creation messages for 'Patient' and 'Biospecimen' manifests
        - manifest csvs and json schemas were created (then removed)

        """
        try:
            # TODO: Set specific paths for csv and json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(manifest, ["--config", "config_example.yml", "get"])

            # command has no (python) errors, has exit code 0
            assert result.exit_code == 0

            biospecimen_df = read_csv("tests/data/example.Biospecimen.manifest.csv")
            patient_df = read_csv("tests/data/example.Patient.manifest.csv")

        # Remove created files:
        finally:
            if os.path.isfile("tests/data/example.Biospecimen.manifest.csv"):
                os.remove("tests/data/example.Biospecimen.manifest.csv")
            if os.path.isfile("tests/data/example.Patient.manifest.csv"):
                os.remove("tests/data/example.Patient.manifest.csv")
            if os.path.isfile("tests/data/example.Biospecimen.schema.json"):
                os.remove("tests/data/example.Biospecimen.schema.json")
            if os.path.isfile("tests/data/example.Patient.schema.json"):
                os.remove("tests/data/example.Patient.schema.json")

        # command output has file creation messages for 'Patient' and 'Biospecimen' manifests
        assert result.output.split("\n")[7] == (
            "Find the manifest template using this CSV file path: "
            "tests/data/example.Biospecimen.manifest.csv"
        )
        assert result.output.split("\n")[10] == (
            "Find the manifest template using this CSV file path: "
            "tests/data/example.Patient.manifest.csv"
        )

        # manifests have expected columns
        assert list(biospecimen_df.columns) == [
            "Sample ID",
            "Patient ID",
            "Tissue Status",
            "Component",
        ]
        assert list(patient_df.columns) == [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Component",
            "Cancer Type",
            "Family History",
        ]
        # manifests only have one row
        assert len(biospecimen_df.index) == 1
        assert len(patient_df.index) == 1
        # manifests are empty except for component column which contains the name of the component
        assert biospecimen_df["Component"].to_list() == ["Biospecimen"]
        assert patient_df["Component"].to_list() == ["Patient"]
        for column in ["Sample ID", "Patient ID", "Tissue Status"]:
            assert np.isnan(biospecimen_df[column].to_list()[0])
        for column in [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Cancer Type",
            "Family History",
        ]:
            assert np.isnan(patient_df[column].to_list()[0])

    @pytest.mark.manual_verification_required
    def test_generate_empty_google_sheet_manifests(
        self,
        runner: CliRunner,
        testing_config: ConfigurationForTesting,
    ) -> None:
        """
        Tests for:
        - command has no errors, has exit code 0
        - command output has file creation messages for 'Patient' and 'Biospecimen' manifest csvs
        - command output has file creation messages for 'Patient' and 'Biospecimen' manifest links

        both google sheets:
        - have drop downs are populated correctly
        - required fields are marked as “light blue”,
        - non-required field are marked as white.
        - first row comments are 'TBD'

        Patient google sheet:
        - first row of 'Family History has its own comment

        Manual tests:
        - Open 'CLI_TestManifestCommand_google_sheet_empty_patient.xlsx'
        - Confirm 'Diagnosis' column to be 'cancer' in the first row
        - Confirm 'Cancer Type' and 'Family History' cells in first row should be light blue.

        """
        try:
            # TODO: Set specific paths json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(
                manifest,
                [
                    "--config",
                    "config_example.yml",
                    "get",
                    "--sheet_url",
                    "--output_csv",
                    "CLI_empty_gs.csv",
                ],
            )
            # command has no errors, has exit code 0
            assert result.exit_code == 0

        finally:
            # Remove created files:
            os.remove("CLI_empty_gs.csv")
            if os.path.isfile("tests/data/example.Biospecimen.schema.json"):
                os.remove("tests/data/example.Biospecimen.schema.json")
            if os.path.isfile("tests/data/example.Patient.schema.json"):
                os.remove("tests/data/example.Patient.schema.json")

        # command output has file creation messages for 'Patient' and 'Biospecimen' manifest csvs
        assert result.output.split("\n")[9] == (
            "Find the manifest template using this CSV file path: " "CLI_empty_gs.csv"
        )
        assert result.output.split("\n")[14] == (
            "Find the manifest template using this CSV file path: " "CLI_empty_gs.csv"
        )

        # command output has file creation messages for 'Patient' and 'Biospecimen' manifest links
        assert result.output.split("\n")[7] == (
            "Find the manifest template using this Google Sheet URL:"
        )
        assert result.output.split("\n")[8].startswith(
            "https://docs.google.com/spreadsheets/d/"
        )
        assert result.output.split("\n")[12] == (
            "Find the manifest template using this Google Sheet URL:"
        )
        assert result.output.split("\n")[13].startswith(
            "https://docs.google.com/spreadsheets/d/"
        )

        # Get the google sheet urls form the message
        google_sheet_url_biospecimen = result.output.split("\n")[8]
        google_sheet_url_patient = result.output.split("\n")[13]

        # Download the Google Sheets content as an Excel file and load into openpyxl
        export_url = f"{google_sheet_url_biospecimen}/export?format=xlsx"
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        assert sheet1[f"{columns['Sample ID']}1"].value == "Sample ID"
        assert sheet1[f"{columns['Sample ID']}2"].value is None

        assert sheet1[f"{columns['Patient ID']}1"].value == "Patient ID"
        assert sheet1[f"{columns['Patient ID']}2"].value is None

        assert sheet1[f"{columns['Tissue Status']}1"].value == "Tissue Status"
        assert sheet1[f"{columns['Tissue Status']}2"].value is None

        assert sheet1[f"{columns['Component']}1"].value == "Component"
        assert sheet1[f"{columns['Component']}2"].value == "Biospecimen"

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['Component']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Sample ID",
            "Patient ID",
            "Tissue Status",
            "Component",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # drop downs are populated correctly
        data_validations = sheet1.data_validations.dataValidation
        tissue_status_validation = None
        for dv in data_validations:
            if f"{columns['Tissue Status']}2" in dv.sqref:
                tissue_status_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"
        assert tissue_status_validation is not None
        assert tissue_status_validation.type == "list"
        assert tissue_status_validation.formula1 == "Sheet2!$C$2:$C$4"

        # required fields are marked as “light blue”, while other non-required fields are marked as white.
        for col in ["Sample ID", "Patient ID", "Tissue Status", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        # Download the Google Sheets content as an Excel file and load into openpyxl
        export_url = f"{google_sheet_url_patient}/export?format=xlsx"
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert sheet1[f"{columns['Patient ID']}1"].value == "Patient ID"
        assert sheet1[f"{columns['Patient ID']}2"].value is None

        assert sheet1[f"{columns['Sex']}1"].value == "Sex"
        assert sheet1[f"{columns['Sex']}2"].value is None

        assert sheet1[f"{columns['Year of Birth']}1"].value == "Year of Birth"
        assert sheet1[f"{columns['Year of Birth']}2"].value is None

        assert sheet1[f"{columns['Diagnosis']}1"].value == "Diagnosis"
        assert sheet1[f"{columns['Diagnosis']}2"].value is None

        assert sheet1[f"{columns['Component']}1"].value == "Component"
        assert sheet1[f"{columns['Component']}2"].value == "Patient"

        assert sheet1[f"{columns['Cancer Type']}1"].value == "Cancer Type"
        assert sheet1[f"{columns['Cancer Type']}2"].value is None

        assert sheet1[f"{columns['Family History']}1"].value == "Family History"
        assert sheet1[f"{columns['Family History']}2"].value is None

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['Family History']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Component",
            "Cancer Type",
            "Family History",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND the comment in "Family History" cell is as expected
        assert (
            sheet1[f"{columns['Family History']}2"].comment.text
            == "Please enter applicable comma-separated items selected from the set of allowable terms for this attribute. See our data standards for allowable terms"
        )

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        sex_validation = None
        diagnosis_validation = None
        cancer_type_validation = None
        for dv in data_validations:
            if f"{columns['Sex']}2" in dv.sqref:
                sex_validation = dv
                continue
            elif f"{columns['Diagnosis']}2" in dv.sqref:
                diagnosis_validation = dv
                continue
            elif f"{columns['Cancer Type']}2" in dv.sqref:
                cancer_type_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert sex_validation is not None
        assert sex_validation.type == "list"
        assert sex_validation.formula1 == "Sheet2!$B$2:$B$4"

        assert diagnosis_validation is not None
        assert diagnosis_validation.type == "list"
        assert diagnosis_validation.formula1 == "Sheet2!$D$2:$D$3"

        assert cancer_type_validation is not None
        assert cancer_type_validation.type == "list"
        assert cancer_type_validation.formula1 == "Sheet2!$F$2:$F$6"

        # AND the fill colors are as expected
        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == WHITE

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['Diagnosis']}2"].value = "Cancer"

        # AND a copy of the Excel file is saved to the test directory for manual verification
        if testing_config.manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    testing_config.manual_test_verification_path,
                    "CLI_TestManifestCommand_google_sheet_empty_patient.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_generate_empty_excel_manifest(
        self, testing_config: ConfigurationForTesting, runner: CliRunner
    ) -> None:
        """
        Tests for:
        - command has no errors, has exit code 0
        - command output has excel file creation message

        Tests for google sheet:
        - drop downs are populated correctly
        - required fields are marked as “light blue”,
          while other non-required fields are marked as white.
        - first row comments are 'TBD'
        - first row of 'Family History has its own comment


        Manual tests:
        - Open 'CLI_TestManifestCommand_excel_empty_patient.xlsx'
        - Confirm 'Diagnosis' column to be 'cancer' in the first row:
        - Confirm 'Cancer Type' and 'Family History' cells in first row should be light blue.
        """
        try:
            # TODO: Set specific paths for csv and json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(
                manifest,
                [
                    "--config",
                    "config_example.yml",
                    "get",
                    "--output_xlsx",
                    "./CLI_empty_excel.xlsx",
                ],
            )

            # command has no errors, has exit code 0
            assert result.exit_code == 0
            workbook = load_workbook("CLI_empty_excel.xlsx")
        finally:
            # Remove created files:
            os.remove("CLI_empty_excel.xlsx")
            if os.path.isfile("tests/data/example.Patient.schema.json"):
                os.remove("tests/data/example.Patient.schema.json")
            if os.path.isfile("tests/data/example.Biospecimen.schema.json"):
                os.remove("tests/data/example.Biospecimen.schema.json")

        # command output has excel file creation message
        result_list = result.output.split("\n")
        assert (
            "Find the manifest template using this Excel file path: ./CLI_empty_excel.xlsx"
            in result_list
        )

        sheet1 = workbook["Sheet1"]
        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert sheet1[f"{columns['Patient ID']}1"].value == "Patient ID"
        assert sheet1[f"{columns['Patient ID']}2"].value is None

        assert sheet1[f"{columns['Sex']}1"].value == "Sex"
        assert sheet1[f"{columns['Sex']}2"].value is None

        assert sheet1[f"{columns['Year of Birth']}1"].value == "Year of Birth"
        assert sheet1[f"{columns['Year of Birth']}2"].value is None

        assert sheet1[f"{columns['Diagnosis']}1"].value == "Diagnosis"
        assert sheet1[f"{columns['Diagnosis']}2"].value is None

        assert sheet1[f"{columns['Component']}1"].value == "Component"
        assert sheet1[f"{columns['Component']}2"].value == "Patient"

        assert sheet1[f"{columns['Cancer Type']}1"].value == "Cancer Type"
        assert sheet1[f"{columns['Cancer Type']}2"].value is None

        assert sheet1[f"{columns['Family History']}1"].value == "Family History"
        assert sheet1[f"{columns['Family History']}2"].value is None

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['Family History']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each cell in the first row has a comment "TBD"
        for col in [
            "Patient ID",
            "Sex",
            "Year of Birth",
            "Diagnosis",
            "Component",
            "Cancer Type",
            "Family History",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND the comment in "Family History" cell is as expected
        assert (
            sheet1[f"{columns['Family History']}2"].comment.text
            == "Please enter applicable comma-separated items selected from the set of allowable terms for this attribute. See our data standards for allowable terms"
        )

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        sex_validation = None
        diagnosis_validation = None
        cancer_type_validation = None
        for dv in data_validations:
            if f"{columns['Sex']}2" in dv.sqref:
                sex_validation = dv
                continue
            elif f"{columns['Diagnosis']}2" in dv.sqref:
                diagnosis_validation = dv
                continue
            elif f"{columns['Cancer Type']}2" in dv.sqref:
                cancer_type_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert sex_validation is not None
        assert sex_validation.type == "list"
        assert sex_validation.formula1 == "Sheet2!$B$2:$B$4"

        assert diagnosis_validation is not None
        assert diagnosis_validation.type == "list"
        assert diagnosis_validation.formula1 == "Sheet2!$D$2:$D$3"

        assert cancer_type_validation is not None
        assert cancer_type_validation.type == "list"
        assert cancer_type_validation.formula1 == "Sheet2!$F$2:$F$6"

        # AND the fill colors are as expected
        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        for col in ["Patient ID", "Sex", "Diagnosis", "Component"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == LIGHT_BLUE

        for col in ["Year of Birth", "Cancer Type", "Family History"]:
            assert sheet1[f"{columns[col]}2"].fill.start_color.index == WHITE

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['Diagnosis']}2"].value = "Cancer"
        # AND a copy of the Excel file is saved to the test directory for manual verification
        if testing_config.manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    testing_config.manual_test_verification_path,
                    "CLI_TestManifestCommand_excel_empty_patient.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_generate_bulk_rna_google_sheet_manifest(
        self, testing_config: ConfigurationForTesting, runner: CliRunner
    ) -> None:
        """
        Tests for:
        - command has no errors, has exit code 0
        - command output has google sheet and csv message

        Tests for google sheet:
        - drop downs are populated correctly
        - required fields are marked as “light blue”,
          while other non-required fields are marked as white.
        - first row comments are 'TBD'


        Manual tests:
        - Open 'CLI_TestManifestCommand_google_sheet_bulk_rna.xlsx'
        - Confirm 'File Format' column to be 'BAM' in the first row
        - Confirm 'Genome Build' cell in first row should be light blue
        - Confirm 'File Format' column to be 'CRAM' in the second row:
        - Confirm 'Genome Build' and 'Genome FASTA' cells in second row should be light blue.
        """
        try:
            # TODO: Set specific paths for json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(
                manifest,
                [
                    "--config",
                    "tests/data/test_configs/CLI_test_config.yml",
                    "get",
                    "--dataset_id",
                    "syn63923432",
                    "--data_type",
                    "BulkRNA-seqAssay",
                    "--sheet_url",
                    "--output_csv",
                    "./CLI_gs_bulk_rna.csv",
                ],
            )
            assert result.exit_code == 0

        finally:
            # Remove created files:
            os.remove("CLI_gs_bulk_rna.csv")
            if os.path.isfile("tests/data/example.BulkRNA-seqAssay.schema.json"):
                os.remove("tests/data/example.BulkRNA-seqAssay.schema.json")

            # TODO: remove with https://sagebionetworks.jira.com/browse/SCHEMATIC-202
            # Reset config to it's default values
            CONFIG.load_config("config_example.yml")

        result_list = result.output.split("\n")
        assert "Find the manifest template using this Google Sheet URL:" in result_list
        assert (
            "Find the manifest template using this CSV file path: "
            "./CLI_gs_bulk_rna.csv"
        ) in result_list
        google_sheet_result = [
            result
            for result in result_list
            if result.startswith("https://docs.google.com/spreadsheets/d/")
        ]
        assert len(google_sheet_result) == 1
        google_sheet_url = google_sheet_result[0]

        # Download the Google Sheets content as an Excel file and load into openpyxl
        export_url = f"{google_sheet_url}/export?format=xlsx"
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert columns["Filename"] is not None
        assert columns["Sample ID"] is not None
        assert columns["File Format"] is not None
        assert columns["Component"] is not None
        assert columns["Genome Build"] is not None
        assert columns["Genome FASTA"] is not None
        assert columns["entityId"] is not None

        assert sheet1[f"{columns['Filename']}2"].value is None
        assert (
            sheet1[f"{columns['Filename']}3"].value
            == "Schematic CLI automation resources/TestDataset1/Sample_A.csv"
        )
        assert (
            sheet1[f"{columns['Filename']}4"].value
            == "Schematic CLI automation resources/TestDataset1/Sample_B.csv"
        )
        assert (
            sheet1[f"{columns['Filename']}5"].value
            == "Schematic CLI automation resources/TestDataset1/Sample_C.csv"
        )
        assert sheet1[f"{columns['Sample ID']}2"].value == 2022
        assert sheet1[f"{columns['Sample ID']}3"].value is None
        assert sheet1[f"{columns['Sample ID']}4"].value is None
        assert sheet1[f"{columns['Sample ID']}5"].value is None
        assert sheet1[f"{columns['File Format']}2"].value == "CSV/TSV"
        assert sheet1[f"{columns['File Format']}3"].value is None
        assert sheet1[f"{columns['File Format']}4"].value is None
        assert sheet1[f"{columns['File Format']}5"].value is None
        assert sheet1[f"{columns['Component']}2"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Component']}3"].value is None
        assert sheet1[f"{columns['Component']}4"].value is None
        assert sheet1[f"{columns['Component']}5"].value is None
        assert sheet1[f"{columns['Genome Build']}2"].value == "GRCm38"
        assert sheet1[f"{columns['Genome Build']}3"].value is None
        assert sheet1[f"{columns['Genome Build']}4"].value is None
        assert sheet1[f"{columns['Genome Build']}5"].value is None
        assert sheet1[f"{columns['Genome FASTA']}2"].value is None
        assert sheet1[f"{columns['Genome FASTA']}3"].value is None
        assert sheet1[f"{columns['Genome FASTA']}4"].value is None
        assert sheet1[f"{columns['Genome FASTA']}5"].value is None
        assert sheet1[f"{columns['entityId']}2"].value == "syn28278954"
        assert sheet1[f"{columns['entityId']}3"].value == "syn63923439"
        assert sheet1[f"{columns['entityId']}4"].value == "syn63923441"
        assert sheet1[f"{columns['entityId']}5"].value == "syn63923444"

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['entityId']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each of these cells in the first row has a comment "TBD"
        for col in [
            "Filename",
            "Sample ID",
            "File Format",
            "Component",
            "Genome Build",
            "Genome FASTA",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND each of these cells in the first row do not have a comment
        for col in [
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].comment is None

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        file_format_validation = None
        genome_build_validation = None
        for dv in data_validations:
            if f"{columns['File Format']}2" in dv.sqref:
                file_format_validation = dv
                continue
            elif f"{columns['Genome Build']}2" in dv.sqref:
                genome_build_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert file_format_validation is not None
        assert file_format_validation.type == "list"
        assert (
            file_format_validation.formula1
            == f"Sheet2!${columns['File Format']}$2:${columns['File Format']}$5"
        )

        assert genome_build_validation is not None
        assert genome_build_validation.type == "list"
        assert (
            genome_build_validation.formula1
            == f"Sheet2!${columns['Genome Build']}$2:${columns['Genome Build']}$5"
        )

        # AND the fill colors are as expected
        for col in ["Filename", "Sample ID", "File Format", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in [
            "Genome Build",
            "Genome FASTA",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['File Format']}2"].value = "BAM"
        workbook["Sheet1"][f"{columns['File Format']}3"].value = "CRAM"
        workbook["Sheet1"][f"{columns['File Format']}4"].value = "FASTQ"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert sheet2["A1"].value == "Filename"
        assert sheet2["B1"].value == "Sample ID"
        assert sheet2["C1"].value == "File Format"
        assert sheet2["D1"].value == "Component"
        assert sheet2["E1"].value == "Genome Build"
        assert sheet2["F1"].value == "Genome FASTA"

        assert sheet2["A2"].value is None
        assert sheet2["B2"].value is None
        assert sheet2["C2"].value == "BAM"
        assert sheet2["D2"].value is None
        assert sheet2["E2"].value == "GRCh37"
        assert sheet2["F2"].value is None

        assert sheet2["A3"].value is None
        assert sheet2["B3"].value is None
        assert sheet2["C3"].value == "CRAM"
        assert sheet2["D3"].value is None
        assert sheet2["E3"].value == "GRCh38"
        assert sheet2["F3"].value is None

        assert sheet2["A4"].value is None
        assert sheet2["B4"].value is None
        assert sheet2["C4"].value == "CSV/TSV"
        assert sheet2["D4"].value is None
        assert sheet2["E4"].value == "GRCm38"
        assert sheet2["F4"].value is None

        assert sheet2["A5"].value is None
        assert sheet2["B5"].value is None
        assert sheet2["C5"].value == "FASTQ"
        assert sheet2["D5"].value is None
        assert sheet2["E5"].value == "GRCm39"
        assert sheet2["F5"].value is None

        # AND there are no more columns in the second sheet
        assert sheet2["G1"].value is None

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['File Format']}2"].value = "BAM"
        workbook["Sheet1"][f"{columns['File Format']}3"].value = "CRAM"
        # A copy of the Excel file is saved to the test directory for manual verification
        if testing_config.manual_test_verification_enabled:
            workbook.save(
                os.path.join(
                    testing_config.manual_test_verification_path,
                    "CLI_TestManifestCommand_google_sheet_bulk_rna.xlsx",
                )
            )

    @pytest.mark.manual_verification_required
    def test_generate_bulk_rna_google_sheet_manifest_with_annotations(
        self, testing_config: ConfigurationForTesting, runner: CliRunner
    ) -> None:
        """
        Tests for:
        - command has no errors, has exit code 0
        - command output has google sheet and csv message

        Tests for google sheet:
        - drop downs are populated correctly
        - required fields are marked as “light blue”,
          while other non-required fields are marked as white.
        - first row comments are 'TBD'


        Manual tests:
        - Open CLI_TestManifestCommand_google_sheet_bulk_rna_with_annotations_url.txt
        - Open the google sheet link in the above file in a browser
        - In the first row  the File Format column should be txt. Hover over it, and there should be an Invalid error.
        - In the second row  the File Format column should be csv. Hover over it, and there should be an Invalid error.
        """
        try:
            # TODO: Set specific paths for json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(
                manifest,
                [
                    "--config",
                    "config_example.yml",
                    "get",
                    "--dataset_id",
                    "syn25614635",
                    "--data_type",
                    "BulkRNA-seqAssay",
                    "--sheet_url",
                    "--use_annotations",
                    "--output_csv",
                    "./CLI_gs_bulk_rna_annos.csv",
                ],
            )
            assert result.exit_code == 0
        finally:
            # Remove created files:
            if os.path.isfile("tests/data/example.BulkRNA-seqAssay.schema.json"):
                os.remove("tests/data/example.BulkRNA-seqAssay.schema.json")
            os.remove("./CLI_gs_bulk_rna_annos.csv")

        result_list = result.output.split("\n")
        assert "Find the manifest template using this Google Sheet URL:" in result_list
        assert (
            "Find the manifest template using this CSV file path: "
            "./CLI_gs_bulk_rna_annos.csv"
        ) in result_list
        google_sheet_result = [
            result
            for result in result_list
            if result.startswith("https://docs.google.com/spreadsheets/d/")
        ]
        assert len(google_sheet_result) == 1
        google_sheet_url = google_sheet_result[0]

        # Download the Google Sheets content as an Excel file and load into openpyxl
        export_url = f"{google_sheet_url}/export?format=xlsx"
        response = requests.get(export_url)
        assert response.status_code == 200
        content = BytesIO(response.content)
        workbook = load_workbook(content)
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]

        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert columns["Filename"] is not None
        assert columns["Sample ID"] is not None
        assert columns["File Format"] is not None
        assert columns["Component"] is not None
        assert columns["Genome Build"] is not None
        assert columns["Genome FASTA"] is not None
        assert columns["impact"] is not None
        assert columns["author"] is not None
        assert columns["eTag"] is not None
        assert columns["IsImportantText"] is not None
        assert columns["IsImportantBool"] is not None
        assert columns["confidence"] is not None
        assert columns["date"] is not None
        assert columns["Year of Birth"] is not None
        assert columns["entityId"] is not None

        assert (
            sheet1[f"{columns['Filename']}2"].value
            == "schematic - main/TestDatasets/TestDataset-Annotations-v3/Sample_A.txt"
        )
        assert sheet1[f"{columns['Sample ID']}2"].value is None
        assert sheet1[f"{columns['File Format']}2"].value == "txt"
        assert sheet1[f"{columns['Component']}2"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}2"].value is None
        assert sheet1[f"{columns['Genome FASTA']}2"].value is None
        assert sheet1[f"{columns['impact']}2"].value is not None
        assert sheet1[f"{columns['author']}2"].value is not None
        assert sheet1[f"{columns['eTag']}2"].value is not None
        assert sheet1[f"{columns['IsImportantText']}2"].value is not None
        assert sheet1[f"{columns['IsImportantBool']}2"].value is not None
        assert sheet1[f"{columns['confidence']}2"].value is not None
        assert sheet1[f"{columns['date']}2"].value is None
        assert sheet1[f"{columns['Year of Birth']}2"].value is not None
        assert sheet1[f"{columns['entityId']}2"].value is not None

        assert (
            sheet1[f"{columns['Filename']}3"].value
            == "schematic - main/TestDatasets/TestDataset-Annotations-v3/Sample_B.txt"
        )
        assert sheet1[f"{columns['Sample ID']}3"].value is None
        assert sheet1[f"{columns['File Format']}3"].value == "csv"
        assert sheet1[f"{columns['Component']}3"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}3"].value is None
        assert sheet1[f"{columns['Genome FASTA']}3"].value is None
        assert sheet1[f"{columns['impact']}3"].value is None
        assert sheet1[f"{columns['author']}3"].value is None
        assert sheet1[f"{columns['eTag']}3"].value is not None
        assert sheet1[f"{columns['IsImportantText']}3"].value is None
        assert sheet1[f"{columns['IsImportantBool']}3"].value is None
        assert sheet1[f"{columns['confidence']}3"].value is not None
        assert sheet1[f"{columns['date']}3"].value is not None
        assert sheet1[f"{columns['Year of Birth']}3"].value is None
        assert sheet1[f"{columns['entityId']}3"].value is not None

        assert (
            sheet1[f"{columns['Filename']}4"].value
            == "schematic - main/TestDatasets/TestDataset-Annotations-v3/Sample_C.txt"
        )
        assert sheet1[f"{columns['Sample ID']}4"].value is None
        assert sheet1[f"{columns['File Format']}4"].value == "fastq"
        assert sheet1[f"{columns['Component']}4"].value == "BulkRNA-seqAssay"
        assert sheet1[f"{columns['Genome Build']}4"].value is None
        assert sheet1[f"{columns['Genome FASTA']}4"].value is None
        assert sheet1[f"{columns['impact']}4"].value is None
        assert sheet1[f"{columns['author']}4"].value is None
        assert sheet1[f"{columns['eTag']}4"].value is not None
        assert sheet1[f"{columns['IsImportantText']}4"].value is not None
        assert sheet1[f"{columns['IsImportantBool']}4"].value is not None
        assert sheet1[f"{columns['confidence']}4"].value is None
        assert sheet1[f"{columns['date']}4"].value is None
        assert sheet1[f"{columns['Year of Birth']}4"].value is None
        assert sheet1[f"{columns['entityId']}4"].value is not None

        # AND there are no more columns in the first sheet
        assert sheet1[f"{columns['entityId']}1"].offset(column=1).value is None

        # AND the first row is locked on scroll
        assert sheet1.freeze_panes == "A2"

        # AND each of these cells in the first row has a comment "TBD"
        for col in [
            "Filename",
            "Sample ID",
            "File Format",
            "Component",
            "Genome Build",
            "Genome FASTA",
        ]:
            assert sheet1[f"{columns[col]}1"].comment.text == "TBD"

        # AND each of these cells in the first row do not have a comment
        for col in [
            "impact",
            "author",
            "eTag",
            "IsImportantText",
            "IsImportantBool",
            "confidence",
            "date",
            "Year of Birth",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].comment is None

        # AND the dropdown lists exist and are as expected
        data_validations = sheet1.data_validations.dataValidation
        file_format_validation = None
        genome_build_validation = None
        for dv in data_validations:
            if f"{columns['File Format']}2" in dv.sqref:
                file_format_validation = dv
                continue
            elif f"{columns['Genome Build']}2" in dv.sqref:
                genome_build_validation = dv
                continue
            # AND there are no other data validations
            assert False, f"Unexpected data validation found: {dv}"

        assert file_format_validation is not None
        assert file_format_validation.type == "list"
        assert (
            file_format_validation.formula1
            == f"Sheet2!${columns['File Format']}$2:${columns['File Format']}$5"
        )

        assert genome_build_validation is not None
        assert genome_build_validation.type == "list"
        assert (
            genome_build_validation.formula1
            == f"Sheet2!${columns['Genome Build']}$2:${columns['Genome Build']}$5"
        )

        # AND the fill colors are as expected
        for col in ["Filename", "Sample ID", "File Format", "Component"]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        for col in [
            "Genome Build",
            "Genome FASTA",
            "impact",
            "author",
            "eTag",
            "IsImportantText",
            "IsImportantBool",
            "confidence",
            "date",
            "Year of Birth",
            "entityId",
        ]:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY

        # AND conditional formatting is functioning as expected (MANUAL VERIFICATION)
        workbook["Sheet1"][f"{columns['File Format']}2"].value = "BAM"
        workbook["Sheet1"][f"{columns['File Format']}3"].value = "CRAM"
        workbook["Sheet1"][f"{columns['File Format']}4"].value = "FASTQ"

        # AND the workbook contains two sheets: "Sheet1" and "Sheet2"
        assert workbook.sheetnames == ["Sheet1", "Sheet2"]

        # AND the second sheet is hidden
        assert sheet2.sheet_state == "hidden"

        # AND the values in "Sheet2" are as expected
        assert sheet2["A1"].value == "Filename"
        assert sheet2["B1"].value == "Sample ID"
        assert sheet2["C1"].value == "File Format"
        assert sheet2["D1"].value == "Component"
        assert sheet2["E1"].value == "Genome Build"
        assert sheet2["F1"].value == "Genome FASTA"

        assert sheet2["A2"].value is None
        assert sheet2["B2"].value is None
        assert sheet2["C2"].value == "BAM"
        assert sheet2["D2"].value is None
        assert sheet2["E2"].value == "GRCh37"
        assert sheet2["F2"].value is None

        assert sheet2["A3"].value is None
        assert sheet2["B3"].value is None
        assert sheet2["C3"].value == "CRAM"
        assert sheet2["D3"].value is None
        assert sheet2["E3"].value == "GRCh38"
        assert sheet2["F3"].value is None

        assert sheet2["A4"].value is None
        assert sheet2["B4"].value is None
        assert sheet2["C4"].value == "CSV/TSV"
        assert sheet2["D4"].value is None
        assert sheet2["E4"].value == "GRCm38"
        assert sheet2["F4"].value is None

        assert sheet2["A5"].value is None
        assert sheet2["B5"].value is None
        assert sheet2["C5"].value == "FASTQ"
        assert sheet2["D5"].value is None
        assert sheet2["E5"].value == "GRCm39"
        assert sheet2["F5"].value is None

        # AND there are no more columns in the second sheet
        assert sheet2["G1"].value is None

        # A copy of the Excel file is saved to the test directory for manual verification
        if testing_config.manual_test_verification_enabled:
            path = os.path.join(
                testing_config.manual_test_verification_path,
                "CLI_TestManifestCommand_google_sheet_bulk_rna_with_annotations_url.txt",
            )
            with open(path, "w") as f:
                f.write(google_sheet_url)

    def test_generate_mock_component_excel_manifest(self, runner: CliRunner) -> None:
        """
        Tests for:
        - Command has no errors, has exit code 0
        - Command output has excel file message
        """
        try:
            # TODO: Set specific paths for json output files with https://sagebionetworks.jira.com/browse/SCHEMATIC-209
            result = runner.invoke(
                manifest,
                [
                    "--config",
                    "tests/data/test_configs/CLI_test_config2.yml",
                    "get",
                    "--output_xlsx",
                    "./CLI_mock_comp.xlsx",
                    "--dataset_id",
                    "syn52746566",
                ],
            )

            # Command has no errors, has exit code 0
            assert result.exit_code == 0
            workbook = load_workbook("./CLI_mock_comp.xlsx")
        finally:
            # Remove created files:
            if os.path.isfile("tests/data/example.MockComponent.schema.json"):
                os.remove("tests/data/example.MockComponent.schema.json")
            os.remove("./CLI_mock_comp.xlsx")

            # TODO: remove with https://sagebionetworks.jira.com/browse/SCHEMATIC-202
            # Reset config to it's default values
            CONFIG.load_config("config_example.yml")
        # Command output has excel file message
        result_list = result.output.split("\n")
        assert (
            "Find the manifest template using this Excel file path: ./CLI_mock_comp.xlsx"
            in result_list
        )

        sheet1 = workbook["Sheet1"]
        # Track column positions
        columns = {cell.value: cell.column_letter for cell in sheet1[1]}

        # AND the content of the first sheet is as expected
        assert sheet1[f"{columns['Component']}1"].value == "Component"
        assert sheet1[f"{columns['Component']}2"].value == "MockComponent"
        assert sheet1[f"{columns['Component']}3"].value == "MockComponent"

        assert sheet1[f"{columns['Check List']}1"].value == "Check List"
        assert sheet1[f"{columns['Check List']}2"].value is not None
        assert sheet1[f"{columns['Check List']}3"].value is not None

        assert sheet1[f"{columns['Check Regex List']}1"].value == "Check Regex List"
        assert sheet1[f"{columns['Check Regex List']}2"].value is not None
        assert sheet1[f"{columns['Check Regex List']}3"].value is not None

        assert sheet1[f"{columns['Check Regex Single']}1"].value == "Check Regex Single"
        assert sheet1[f"{columns['Check Regex Single']}2"].value is not None
        assert sheet1[f"{columns['Check Regex Single']}3"].value is not None

        assert sheet1[f"{columns['Check Regex Format']}1"].value == "Check Regex Format"
        assert sheet1[f"{columns['Check Regex Format']}2"].value is not None
        assert sheet1[f"{columns['Check Regex Format']}3"].value is not None

        assert (
            sheet1[f"{columns['Check Regex Integer']}1"].value == "Check Regex Integer"
        )
        assert sheet1[f"{columns['Check Regex Integer']}2"].value is not None
        assert sheet1[f"{columns['Check Regex Integer']}3"].value is not None

        assert sheet1[f"{columns['Check Num']}1"].value == "Check Num"
        assert sheet1[f"{columns['Check Num']}2"].value is not None
        assert sheet1[f"{columns['Check Num']}3"].value is not None

        assert sheet1[f"{columns['Check Float']}1"].value == "Check Float"
        assert sheet1[f"{columns['Check Float']}2"].value is not None
        assert sheet1[f"{columns['Check Float']}3"].value is not None

        assert sheet1[f"{columns['Check Int']}1"].value == "Check Int"
        assert sheet1[f"{columns['Check Int']}2"].value is not None
        assert sheet1[f"{columns['Check Int']}3"].value is not None

        required_columns = [
            "Component",
            "Check List",
            "Check List Enum",
            "Check List Like",
            "Check List Like Enum",
            "Check List Strict",
            "Check List Enum Strict",
            "Check Regex List",
            "Check Regex List Strict",
            "Check Regex List Like",
            "Check Regex Single",
            "Check Regex Format",
            "Check Regex Integer",
            "Check Num",
            "Check Float",
            "Check Int",
            "Check String",
            "Check URL",
            "Check Match at Least",
            "Check Match Exactly",
            "Check Match None",
            "Check Match at Least values",
            "Check Match Exactly values",
            "Check Match None values",
            "Check Ages",
            "Check Unique",
            "Check Range",
            "Check Date",
            "Check NA",
        ]

        optional_columns = [
            "Patient ID",
            "Sex",
            "Diagnosis",
            "Cancer Type",
            "Family History",
            "Sample ID",
            "Tissue Status",
            "Filename",
            "File Format",
            "Genome Build",
            "Genome FASTA",
            "Patient",
            "Year of Birth",
            "Cancer",
            "Biospecimen",
            "Bulk RNA-seq Assay",
            "BAM",
            "CRAM",
            "CSV/TSV",
            "MockComponent",
            "Check Recommended",
            "MockRDB",
            "MockFilename",
            "MockRDB_id",
            "SourceManifest",
        ]

        # Required columns are light blue
        for col in required_columns:
            assert sheet1[f"{columns[col]}1"].fill.start_color.index == LIGHT_BLUE

        # Optional columns are in grey
        for col in optional_columns:
            if col in columns:
                assert sheet1[f"{columns[col]}1"].fill.start_color.index == GRAY


class TestDownloadManifest:
    """Tests the command line interface for downloading a manifest"""

    def test_download_manifest_found(
        self,
        runner: CliRunner,
        config: Configuration,
    ) -> None:
        # GIVEN a manifest name to download as
        manifest_name = f"{uuid.uuid4()}"

        # AND a dataset id
        dataset_id = "syn23643250"

        # AND a configuration file
        config.load_config("config_example.yml")

        # WHEN the download command is run
        result = runner.invoke(
            cli=manifest,
            args=[
                "--config",
                config.config_path,
                "download",
                "--new_manifest_name",
                manifest_name,
                "--dataset_id",
                dataset_id,
            ],
        )

        # THEN the command should run successfully
        assert result.exit_code == 0

        # AND the manifest file should be created
        expected_manifest_file = os.path.join(
            config.manifest_folder, f"{manifest_name}.csv"
        )
        assert os.path.exists(expected_manifest_file)
        try:
            os.remove(expected_manifest_file)
        except Exception:
            pass

    def test_download_manifest_not_found(
        self,
        runner: CliRunner,
        config: Configuration,
    ) -> None:
        # GIVEN a manifest name to download as
        manifest_name = f"{uuid.uuid4()}"

        # AND a dataset id that does not exist
        dataset_id = "syn1234"

        # AND a configuration file
        config.load_config("config_example.yml")

        # WHEN the download command is run
        result = runner.invoke(
            cli=manifest,
            args=[
                "--config",
                config.config_path,
                "download",
                "--new_manifest_name",
                manifest_name,
                "--dataset_id",
                dataset_id,
            ],
        )

        # THEN the command should not run successfully
        assert result.exit_code == 1

        # AND the manifest file should not be created
        expected_manifest_file = os.path.join(
            config.manifest_folder, f"{manifest_name}.csv"
        )
        assert not os.path.exists(expected_manifest_file)
