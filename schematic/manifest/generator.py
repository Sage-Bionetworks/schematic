import json
import logging
import os
from collections import OrderedDict
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, BinaryIO, Dict, List, Literal, Optional, Tuple, Union

import networkx as nx
import pandas as pd
import pygsheets as ps
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from opentelemetry import trace

from schematic.configuration.configuration import CONFIG
from schematic.schemas.data_model_graph import DataModelGraph, DataModelGraphExplorer
from schematic.schemas.data_model_json_schema import DataModelJSONSchema
from schematic.schemas.data_model_parser import DataModelParser

# TODO: This module should only be aware of the store interface
# we shouldn't need to expose Synapse functionality explicitly
from schematic.store.synapse import SynapseStorage
from schematic.utils.df_utils import load_df, update_df
from schematic.utils.google_api_utils import (
    build_service_account_creds,
    execute_google_api_requests,
    export_manifest_drive_service,
)
from schematic.utils.schema_utils import (
    DisplayLabelType,
    extract_component_validation_rules,
)
from schematic.utils.validate_utils import rule_in_rule_list

logger = logging.getLogger(__name__)
tracer = trace.get_tracer("Schematic")


class ManifestGenerator(object):
    def __init__(
        self,
        path_to_data_model: str,  # JSON-LD file to be used for generating the manifest
        graph: nx.MultiDiGraph,  # At this point, the graph is fully formed.
        alphabetize_valid_values: str = "ascending",
        title: str = None,  # manifest sheet title
        root: str = None,
        additional_metadata: Dict = None,
        use_annotations: bool = False,
    ) -> None:
        # use service account creds
        services_creds = build_service_account_creds()

        # google service for Sheet API
        self.sheet_service = services_creds["sheet_service"]

        # google service for Drive API
        self.drive_service = services_creds["drive_service"]

        # google service credentials object
        self.creds = services_creds["creds"]

        # Path to jsonld
        self.model_path = path_to_data_model

        # Graph
        self.graph = graph

        # schema root
        if root:
            self.root = root
        # Raise an error if no DataType has been provided
        else:
            raise ValueError("No DataType has been provided.")

        # alphabetize valid values
        self.alphabetize = alphabetize_valid_values

        # manifest title
        self.title = title
        if self.title is None:
            self.title = f"{self.root} - Manifest"

        # Whether to use existing annotations during manifest generation
        self.use_annotations = use_annotations

        # Warn about limited feature support for `use_annotations`
        if self.use_annotations:
            logger.warning(
                "The `use_annotations` option is currently only supported "
                "when there is no manifest file for the dataset in question."
            )

        # Instantiate Data Model Explorer object
        self.dmge = DataModelGraphExplorer(self.graph)

        # additional metadata to add to manifest
        self.additional_metadata = additional_metadata

        # Check if the class is in the schema
        root_in_schema = self.dmge.is_class_in_schema(self.root)

        # If the class could not be found, give a notification
        if not root_in_schema:
            exception_message = (
                f"The DataType entered ({self.root}) could not be found in the data model schema. "
                + "Please confirm that the datatype is in the data model and that the spelling matches the class label in the .jsonld file."
            )
            raise LookupError(exception_message)

        # Determine whether current data type is file-based
        self.is_file_based = "Filename" in self.dmge.get_node_dependencies(self.root)

    def _attribute_to_letter(self, attribute, manifest_fields):
        """Map attribute to column letter in a google sheet"""

        # find index of attribute in manifest field
        column_idx = manifest_fields.index(attribute)

        # return the google sheet letter representation of the column index
        return self._column_to_letter(column_idx)

    def _column_to_letter(self, column):
        """Find google sheet letter representation of a column index integer"""
        character = chr(ord("A") + column % 26)
        remainder = column // 26
        if column >= 26:
            return self._column_to_letter(remainder - 1) + character
        else:
            return character

    def _columns_to_sheet_ranges(self, column_idxs):
        """map a set of column indexes to a set of Google sheet API ranges: each range includes exactly one column"""
        ranges = []

        for column_idx in column_idxs:
            col_range = {
                "startColumnIndex": column_idx,
                "endColumnIndex": column_idx + 1,
            }

            ranges.append(col_range)

        return ranges

    def _column_to_cond_format_eq_rule(
        self, column_idx: int, condition_argument: str, required: bool = False
    ) -> dict:
        """Given a column index and an equality argument (e.g. one of valid values for the given column fields), generate a conditional formatting rule based on a custom formula encoding the logic:

        'if a cell in column idx is equal to condition argument, then set specified formatting'
        """

        col_letter = self._column_to_letter(column_idx)

        if not required:
            bg_color = CONFIG.google_optional_background_color
        else:
            bg_color = CONFIG.google_required_background_color

        boolean_rule = {
            "condition": {
                "type": "CUSTOM_FORMULA",
                "values": [
                    {
                        "userEnteredValue": "=$"
                        + col_letter
                        + '1 = "'
                        + condition_argument
                        + '"'
                    }
                ],
            },
            "format": {"backgroundColor": bg_color},
        }

        return boolean_rule

    def _gdrive_copy_file(self, origin_file_id, copy_title):
        """Copy an existing file.

        Args:
            origin_file_id: ID of the origin file to copy.
            copy_title: Title of the copy.

        Returns:
            The copied file if successful, None otherwise.
        """
        copied_file = {"name": copy_title}

        # return new copy sheet ID
        return (
            self.drive_service.files()
            .copy(fileId=origin_file_id, body=copied_file)
            .execute()["id"]
        )

    def _create_empty_manifest_spreadsheet(self, title: str) -> str:
        """
        Creates an empty google spreadsheet returning the id.
        If the configuration has a template id it will be used

        Args:
            title (str): The title of the spreadsheet

        Returns:
            str: The id of the created spreadsheet
        """
        template_id = CONFIG.google_sheets_master_template_id

        if template_id:
            spreadsheet_id = self._gdrive_copy_file(template_id, title)

        else:
            spreadsheet_body = {"properties": {"title": title}}

            spreadsheet_id = (
                self.sheet_service.spreadsheets()
                .create(body=spreadsheet_body, fields="spreadsheetId")
                .execute()
                .get("spreadsheetId")
            )

        return spreadsheet_id

    def _get_cell_borders(self, cell_range):
        # set border style request
        color = {
            "red": 226.0 / 255.0,
            "green": 227.0 / 255.0,
            "blue": 227.0 / 255.0,
        }

        border_style_req = {
            "updateBorders": {
                "range": cell_range,
                "top": {"style": "SOLID", "width": 2, "color": color},
                "bottom": {"style": "SOLID", "width": 2, "color": color},
                "left": {"style": "SOLID", "width": 2, "color": color},
                "right": {"style": "SOLID", "width": 2, "color": color},
                "innerHorizontal": {"style": "SOLID", "width": 2, "color": color},
                "innerVertical": {"style": "SOLID", "width": 2, "color": color},
            }
        }

        return border_style_req

    def _set_permissions(self, fileId):
        def callback(request_id, response, exception):
            if exception:
                # Handle error
                logger.error(exception)
            else:
                logger.info(f"Permission Id: {response.get('id')}")

        batch = self.drive_service.new_batch_http_request(callback=callback)

        worldPermission = {"type": "anyone", "role": "writer"}

        batch.add(
            self.drive_service.permissions().create(
                fileId=fileId,
                body=worldPermission,
                fields="id",
            )
        )
        batch.execute()

    def _store_valid_values_as_data_dictionary(
        self, column_id: int, valid_values: list, spreadsheet_id: str
    ) -> list:
        """store valid values in google sheet (sheet 2). This step is required for "ONE OF RANGE" validation
        Args:
            column_id: id of column
            valid_values: a list of valid values for a given attribute (i.e. for diagnosis, this looks like: [{'userEnteredValue': 'Cancer'}, {'userEnteredValue': 'Healthy'}])
            spreadsheet_id: google spreadsheet id

        return: range of valid values (i.e. for diagnosis, [{'userEnteredValue': '=Sheet2!D2:D3'}])
        """
        # get valid values w/o google sheet header
        values = [valid_value["userEnteredValue"] for valid_value in valid_values]

        if self.alphabetize and self.alphabetize.lower().startswith("a"):
            values.sort(reverse=False, key=str.lower)
        elif self.alphabetize and self.alphabetize.lower().startswith("d"):
            values.sort(reverse=True, key=str.lower)

        # store valid values explicitly in workbook at the provided range to use as validation values
        target_col_letter = self._column_to_letter(column_id)
        body = {"majorDimension": "COLUMNS", "values": [values]}
        target_range = (
            "Sheet2!"
            + target_col_letter
            + "2:"
            + target_col_letter
            + str(len(values) + 1)
        )
        valid_values = [{"userEnteredValue": "=" + target_range}]
        response = (
            self.sheet_service.spreadsheets()
            .values()
            .update(
                spreadsheetId=spreadsheet_id,
                range=target_range,
                valueInputOption="RAW",
                body=body,
            )
            .execute()
        )
        return valid_values

    def _get_column_data_validation_values(
        self,
        spreadsheet_id,
        valid_values,
        column_id,
        strict: Optional[bool],
        validation_type="ONE_OF_LIST",
        custom_ui=True,
        input_message="Choose one from dropdown",
    ):
        # set validation strictness to config file default if None indicated.
        if strict == None:
            strict = CONFIG.google_sheets_strict_validation
        # store valid values explicitly in workbook at the provided range to use as validation values
        if validation_type == "ONE_OF_RANGE":
            valid_values = self._store_valid_values_as_data_dictionary(
                column_id, valid_values, spreadsheet_id
            )

        # setup validation data request body
        validation_body = {
            "requests": [
                {
                    "setDataValidation": {
                        "range": {
                            "startRowIndex": 1,
                            "startColumnIndex": column_id,
                            "endColumnIndex": column_id + 1,
                        },
                        "rule": {
                            "condition": {
                                "type": validation_type,
                                "values": valid_values,
                            },
                            "inputMessage": input_message,
                            "strict": strict,
                            "showCustomUi": custom_ui,
                        },
                    }
                }
            ]
        }

        return validation_body

    def _get_valid_values_from_jsonschema_property(self, prop: dict) -> List[str]:
        """Get valid values for a manifest attribute based on the corresponding
        values of node's properties in JSONSchema

        Args:
            prop: node properties - jsonschema dictionary

        Returns:
            List of valid values
        """

        if "enum" in prop:
            return prop["enum"]
        elif "items" in prop:
            return prop["items"]["enum"]
        else:
            return []

    def _get_json_schema(self, json_schema_filepath: str) -> Dict:
        """Open json schema as a dictionary.
        Args:
            json_schema_filepath(str): path to json schema file
        Returns:
            Dictionary, containing portions of the json schema
        TODO: Do we even allow people to provide a json_schema_filepath anyore?
        """
        if not json_schema_filepath:
            # TODO Catch error if no JSONLD or JSON path provided.
            data_model_js = DataModelJSONSchema(
                jsonld_path=self.model_path, graph=self.graph
            )
            json_schema = data_model_js.get_json_validation_schema(
                source_node=self.root, schema_name=self.title
            )
        else:
            with open(json_schema_filepath) as jsonfile:
                json_schema = json.load(jsonfile)
        return json_schema

    def _get_required_metadata_fields(self, json_schema, fields):
        """For the root node gather dependency requirements (all attributes linked to this node)
        and corresponding allowed values constraints (i.e. valid values).

        Args:
            json_schema(dict): representing a handful of values
                representing the data model, including: '$schema', '$id', 'title',
                'type', 'properties', 'required'
            fields(list[str]): fields/attributes to search
        Returns:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        """
        required_metadata_fields = {}
        for req in fields:
            required_metadata_fields[
                req
            ] = self._get_valid_values_from_jsonschema_property(
                json_schema["properties"][req]
            )
            # the following line may not be needed
            json_schema["properties"][req]["enum"] = required_metadata_fields[req]
        return required_metadata_fields

    def _gather_dependency_requirements(self, json_schema, required_metadata_fields):
        """Gathering dependency requirements and allowed value constraints
            for conditional dependencies, if any
        TODO:
            Refactor
        Args:
            json_schema(dict): representing a handful of values
                representing the data model, including: '$schema', '$id', 'title',
                'type', 'properties', 'required'
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        Returns:
            required_metadata_fields(dict):
                updates dictionary to additional attributes, if applicable,
                as keys with their corresponding valid values.
        """

        if "allOf" in json_schema:
            for conditional_reqs in json_schema["allOf"]:
                if "required" in conditional_reqs["if"]:
                    for req in conditional_reqs["if"]["required"]:
                        if req in conditional_reqs["if"]["properties"]:
                            if not req in required_metadata_fields:
                                if req in json_schema["properties"]:
                                    required_metadata_fields[
                                        req
                                    ] = self._get_valid_values_from_jsonschema_property(
                                        json_schema["properties"][req]
                                    )
                                else:
                                    required_metadata_fields[
                                        req
                                    ] = self._get_valid_values_from_jsonschema_property(
                                        conditional_reqs["if"]["properties"][req]
                                    )

                    for req in conditional_reqs["then"]["required"]:
                        if not req in required_metadata_fields:
                            if req in json_schema["properties"]:
                                required_metadata_fields[
                                    req
                                ] = self._get_valid_values_from_jsonschema_property(
                                    json_schema["properties"][req]
                                )

        return required_metadata_fields

    def _add_root_to_component(self, required_metadata_fields: Dict[str, List]):
        """If 'Component' is in the column set, add root node as a
        metadata component entry in the first row of that column.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        Return:
            updates self.additional_metadata if appropriate to
            contain {'Component': [self.root]}
        """
        if "Component" in required_metadata_fields.keys():
            # check if additional metadata has actually been instantiated in the
            # constructor (it's optional) if not, instantiate it
            if not self.additional_metadata:
                self.additional_metadata = {}
            if self.is_file_based and "Filename" in self.additional_metadata:
                self.additional_metadata["Component"] = [self.root] * max(
                    1, len(self.additional_metadata["Filename"])
                )
            else:
                self.additional_metadata["Component"] = [self.root]
        return

    def _get_additional_metadata(self, required_metadata_fields: dict) -> dict:
        """Add additional metadata as entries to columns.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        Return:
            required_metadata_fields(dict): Updated required_metadata_fields to include additional metadata
                that is to be added to specified columns
        """
        self._add_root_to_component(required_metadata_fields)
        # if additional metadata is provided append columns (if those do not exist already)
        if self.additional_metadata:
            for column in self.additional_metadata.keys():
                if not column in required_metadata_fields:
                    required_metadata_fields[column] = []
        return required_metadata_fields

    def _get_column_range_and_order(self, required_metadata_fields):
        """Find the alphabetical range of columns and sort them.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        Returns:
            end_col_letter (chr): google sheet letter representation of a column index integer
            ordered_metadata_fields(List[list]): Contining all the attributes to
            add as columns, ordered
        """
        # determining columns range
        end_col = len(required_metadata_fields.keys())
        end_col_letter = self._column_to_letter(end_col)

        # order columns header (since they are generated based on a json schema, which is a dict)
        ordered_metadata_fields = [list(required_metadata_fields.keys())]
        ordered_metadata_fields[0] = self.sort_manifest_fields(
            ordered_metadata_fields[0]
        )
        return end_col_letter, ordered_metadata_fields

    def _gs_add_and_format_columns(self, required_metadata_fields, spreadsheet_id):
        """Add columns to the google sheet and format them.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
            spreadsheet_id(str): id of the google sheet
        Returns:
            respones: In return for a request, used by google sheet api to add columns
                to the sheet being generated.
            ordered_metadata_fields(List[list]): contining all the attributes to
            add as columns, ordered
        """

        end_col_letter, ordered_metadata_fields = self._get_column_range_and_order(
            required_metadata_fields
        )

        body = {"values": ordered_metadata_fields}

        range = "Sheet1!A1:" + str(end_col_letter) + "1"

        # adding columns
        self.sheet_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range, valueInputOption="RAW", body=body
        ).execute()

        # adding columns to 2nd sheet that can be used for storing data validation ranges (this avoids limitations on number of dropdown items in excel and openoffice)
        range = "Sheet2!A1:" + str(end_col_letter) + "1"
        self.sheet_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=range, valueInputOption="RAW", body=body
        ).execute()

        # format column header row
        header_format_body = {
            "requests": [
                {
                    "repeatCell": {
                        "range": {"startRowIndex": 0, "endRowIndex": 1},
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {
                                    "red": 224.0 / 255,
                                    "green": 224.0 / 255,
                                    "blue": 224.0 / 255,
                                },
                                "horizontalAlignment": "CENTER",
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 0.0 / 255,
                                        "green": 0.0 / 255,
                                        "blue": 0.0 / 255,
                                    },
                                    "fontSize": 8,
                                    "bold": True,
                                },
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)",
                    }
                },
                {
                    "updateSheetProperties": {
                        "properties": {"gridProperties": {"frozenRowCount": 1}},
                        "fields": "gridProperties.frozenRowCount",
                    }
                },
                {
                    "autoResizeDimensions": {
                        "dimensions": {"dimension": "COLUMNS", "startIndex": 0}
                    }
                },
            ]
        }

        response = (
            self.sheet_service.spreadsheets()
            .batchUpdate(spreadsheetId=spreadsheet_id, body=header_format_body)
            .execute()
        )
        return response, ordered_metadata_fields

    def _gs_add_additional_metadata(
        self, required_metadata_fields, ordered_metadata_fields, spreadsheet_id
    ):
        """Adding additional metadata values, if needed. Add value-constraints
        from data model as dropdowns. Batch google API request to
        create metadata template.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
            ordered_metadata_fields(List[list]): contining all the attributes to
            add as columns, ordered
            spreadsheet_id(str): id for the google sheet
        Returns: Response(dict): For batch update to create metadata
        template.
        """
        data = []

        for i, req in enumerate(ordered_metadata_fields[0]):
            values = required_metadata_fields[req]

            if self.additional_metadata and req in self.additional_metadata:
                values = self.additional_metadata[req]
                target_col_letter = self._column_to_letter(i)

                range_vals = (
                    target_col_letter + "2:" + target_col_letter + str(len(values) + 1)
                )

                data.append(
                    {
                        "range": range_vals,
                        "majorDimension": "COLUMNS",
                        "values": [values],
                    }
                )

        batch_update_values_request_body = {
            # How the input data should be interpreted.
            "valueInputOption": "RAW",
            # The new values to apply to the spreadsheet.
            "data": data,
        }

        response = (
            self.sheet_service.spreadsheets()
            .values()
            .batchUpdate(
                spreadsheetId=spreadsheet_id, body=batch_update_values_request_body
            )
            .execute()
        )
        return response

    def _request_update_base_color(self, i: int, color={"red": 1.0}):
        """
        Change color of text in column we are validating
        to red.
        """
        vr_format_body = {
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "startColumnIndex": i,
                            "endColumnIndex": i + 1,
                            "startRowIndex": 1,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {"foregroundColor": color}
                            }
                        },
                        "fields": "userEnteredFormat(textFormat)",
                    }
                }
            ]
        }
        return vr_format_body

    def _request_regex_vr(self, gs_formula, i: int, text_color={"red": 1}):
        """
        Generate request to change font color to black upon corretly formatted
        user entry.
        """
        requests_vr = {
            "requests": [
                {
                    "addConditionalFormatRule": {
                        "rule": {
                            "ranges": {
                                "startColumnIndex": i,
                                "endColumnIndex": i + 1,
                                "startRowIndex": 1,
                            },
                            "booleanRule": {
                                "condition": {
                                    "type": "CUSTOM_FORMULA",
                                    "values": gs_formula,
                                },
                                "format": {
                                    "textFormat": {"foregroundColor": text_color}
                                },
                            },
                        },
                        "index": 0,
                    }
                }
            ]
        }
        return requests_vr

    def _request_regex_match_vr_formatting(
        self,
        validation_rules: List[str],
        i: int,
        spreadsheet_id: str,
        requests_body: dict,
        strict: Optional[bool],
    ):
        """
        Purpose:
            - Apply regular expression validaiton rules to google sheets.
            - Will only be run if the regex module specified in the validation
            rules is 'match'.
                - This is because of the limitations of google sheets regex.
                - Users can additionally add 'strict' to the end of their validation
                rules. This would allow the rule itself to set the level of strictness,
                otherwise it would fall to the default level set in the config file.
            - Will do the following:
                - In google sheets user entry text will initially appear red.
                - Upon correct format entry, text will turn black.
                - If incorrect format is entered a validation error will pop up.
        Input:
            validation_rules: List[str], defines the validation rules
                applied to a particular column.
            i: int, defines current column.
            requests_body: dict, containing all the update requests to add to the gs
        Returns:
            requests_body: updated with additional formatting requests if regex match
                is specified.
        """
        split_rules = validation_rules[0].split(" ")
        if split_rules[0] == "regex" and split_rules[1] == "match":
            # Set things up:
            ## Extract the regular expression we are validating against.
            regular_expression = split_rules[2]
            ## Define text color to update to upon correct user entry
            text_color = {"red": 0, "green": 0, "blue": 0}
            ## Define google sheets regular expression formula
            gs_formula = [
                {
                    "userEnteredValue": '=REGEXMATCH(INDIRECT("RC",FALSE), "{}")'.format(
                        regular_expression
                    )
                }
            ]
            ## Set validaiton strictness based on user specifications.
            if split_rules[-1].lower() == "strict":
                strict = True

            ## Create error message for users if they enter value with incorrect formatting
            input_message = (
                f"Values in this column are being validated "
                f"against the following regular expression ({regular_expression}) "
                f"to ensure for accuracy. Please re-enter value according to these "
                f"formatting rules"
            )

            # Create Requests:
            ## Change request to change the text color of the column we are validating to red.
            requests_vr_format_body = self._request_update_base_color(
                i,
                color={
                    "red": 232.0 / 255.0,
                    "green": 80.0 / 255.0,
                    "blue": 70.0 / 255.0,
                },
            )

            ## Create request to for conditionally formatting user input.
            requests_vr = self._request_regex_vr(gs_formula, i, text_color)

            ## Create request to generate data validator.
            requests_data_validation_vr = self._get_column_data_validation_values(
                spreadsheet_id,
                valid_values=gs_formula,
                column_id=i,
                strict=strict,
                custom_ui=False,
                input_message=input_message,
                validation_type="CUSTOM_FORMULA",
            )

            requests_body["requests"].append(requests_vr_format_body["requests"])
            requests_body["requests"].append(requests_vr["requests"])
            requests_body["requests"].append(requests_data_validation_vr["requests"])
        return requests_body

    def _request_row_format(self, i, req):
        """Adding description to headers, this is not executed if
        only JSON schema is defined. Also formatting required columns.
        Args:
            i (int): column index
            req (str): column name
        Returns:
            notes_body["requests"] (dict): with information on note
                to add to the column header. This notes body will be added to a request.
        """
        if self.dmge:
            # get node definition
            note = self.dmge.get_node_comment(node_display_name=req)

            notes_body = {
                "requests": [
                    {
                        "updateCells": {
                            "range": {
                                "startRowIndex": 0,
                                "endRowIndex": 1,
                                "startColumnIndex": i,
                                "endColumnIndex": i + 1,
                            },
                            "rows": [{"values": [{"note": note}]}],
                            "fields": "note",
                        }
                    }
                ]
            }

            return notes_body["requests"]
        else:
            return

    def _request_note_valid_values(self, i, req, validation_rules, valid_values):
        """If the node validation rule is 'list' and the node also contains
        valid values, add a note a note with instructions on adding a list of
        multiple values with the multi-select option in google sheets

        TODO: add validation and QC rules "compiler/generator" class elsewhere
            for now have the list logic here
        Args:
            i (int): column index
            req (str): column name
            validation_rules(list[str]): containing all relevant
                rules applied to node/column
            valid_values(list[str]): containing all valid values defined in the
                data model for this node/column

        Returns:
            notes_body["requests"] (dict): with information on note
                to add to the column header, about using multiselect.
                This notes body will be added to a request.
        """
        if rule_in_rule_list("list", validation_rules) and valid_values:
            note = "Please enter applicable comma-separated items selected from the set of allowable terms for this attribute. See our data standards for allowable terms"
            notes_body = {
                "requests": [
                    {
                        "repeatCell": {
                            "range": {
                                "startRowIndex": 1,
                                "startColumnIndex": i,
                                "endColumnIndex": i + 1,
                            },
                            "cell": {"note": note},
                            "fields": "note",
                        }
                    }
                ]
            }
            return notes_body["requests"]
        elif rule_in_rule_list("list", validation_rules) and not valid_values:
            note = (
                "Please enter values as a comma separated list. For example: XX, YY, ZZ"
            )
            notes_body = {
                "requests": [
                    {
                        "repeatCell": {
                            "range": {
                                "startRowIndex": 1,
                                "startColumnIndex": i,
                                "endColumnIndex": i + 1,
                            },
                            "cell": {"note": note},
                            "fields": "note",
                        }
                    }
                ]
            }
            return notes_body["requests"]
        else:
            return

    def _set_required_columns_color(self, i, req, json_schema):
        """Update background colors so that columns that are required are highlighted
        Args:
            i (int): column index
            req (str): column name
            json_schema(dict): representing a handful of values
                representing the data model, including: '$schema', '$id', 'title',
                'type', 'properties', 'required'
        Returns:
            req_format_body["requests"] (dict): specifing the format updating for
                required attributes/columns
        """
        # check if attribute is required and set a corresponding color
        if req in json_schema["required"]:
            bg_color = CONFIG.google_required_background_color

            req_format_body = {
                "requests": [
                    {
                        "repeatCell": {
                            "range": {
                                "startColumnIndex": i,
                                "endColumnIndex": i + 1,
                            },
                            "cell": {
                                "userEnteredFormat": {"backgroundColor": bg_color}
                            },
                            "fields": "userEnteredFormat(backgroundColor)",
                        }
                    }
                ]
            }
            return req_format_body["requests"]
        else:
            return

    def _request_cell_borders(self):
        """Get cell border color specifications.
        Args: None
        Returns:
            Dict, containing cell border design specifications.
        """
        # setting cell borders
        cell_range = {
            "sheetId": 0,
            "startRowIndex": 0,
        }
        return self._get_cell_borders(cell_range)

    def _request_dropdown(
        self, i, req_vals, spreadsheet_id, validation_rules, valid_values
    ):
        """Generating sheet api request to populate a dropdown
        Note: multi-select was deprecated
        Args:
            i (int): column index
            req_vals (list[dict]): dict for each valid value
                key: 'userEnteredValue'
                value: str, valid value
            spreadsheet_id (str): id for the google sheet
            validation_rules(list[str]): containing all relevant
                rules applied to node/column
            valid_values(list[str]): containing all valid values defined in the
                data model for this node/column
        Returns:
            validation_body: dict
        """
        if len(req_vals) > 0:
            # if more than 0 values in dropdown use ONE_OF_RANGE type of validation
            # since excel and openoffice
            # do not support other kinds of data validation for
            # larger number of items (even if individual items are not that many
            # excel has a total number of characters limit per dropdown...)
            validation_body = self._get_column_data_validation_values(
                spreadsheet_id, req_vals, i, strict=None, validation_type="ONE_OF_RANGE"
            )
        else:
            validation_body = self._get_column_data_validation_values(
                spreadsheet_id, req_vals, i, strict=None
            )
        return validation_body["requests"]

    def _dependency_formatting(
        self,
        i,
        req_val,
        ordered_metadata_fields,
        val_dependencies,
        dependency_formatting_body,
    ):
        """If there are additional attribute dependencies find the corresponding
        fields that need to be filled in and construct conditional formatting rules
        indicating the dependencies need to be filled in.

        set target ranges for this rule
        i.e. dependency attribute columns that will be formatted

        Args:
            i (int): column index
            req_val (str): node name
            ordered_metadata_fields(List[list]): contining all the attributes to
            add as columns, ordered
            val_depenencies (list[str]): dependencies
        Returns:
            dependency_formatting_body["requests"] (list):
                specifies gs conditional formatting per val_dependency
        """

        # find dependency column indexes
        # note that dependencies values must be in index
        # TODO: catch value error that shouldn't happen
        column_idxs = [
            ordered_metadata_fields[0].index(val_dep) for val_dep in val_dependencies
        ]

        # construct ranges based on dependency column indexes
        rule_ranges = self._columns_to_sheet_ranges(column_idxs)
        # go over valid value dependencies
        dependency_formatting_body = {"requests": []}
        for j, val_dep in enumerate(val_dependencies):
            is_required = False
            if self.dmge.get_node_required(node_display_name=val_dep):
                is_required = True
            else:
                is_required = False

            # construct formatting rule
            formatting_rule = self._column_to_cond_format_eq_rule(
                i, req_val, required=is_required
            )

            # construct conditional format rule
            conditional_format_rule = {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": rule_ranges[j],
                        "booleanRule": formatting_rule,
                    },
                    "index": 0,
                }
            }
            dependency_formatting_body["requests"].append(conditional_format_rule)
        return dependency_formatting_body["requests"]

    def _request_dependency_formatting(
        self, i, req_vals, ordered_metadata_fields, requests_body
    ):
        """Gather all the formattng for node dependencies.
        Args:
            i (int): column index
            req_val (str): node name
            ordered_metadata_fields(List[list]): contining all the attributes to
            add as columns, ordered
            requests_body(dict):
                containing all the update requests to add to the gs
        Return:
            requests_body (dict): adding the conditional
            formatting rules to apply
        """
        for req_val in req_vals:
            # get this required/valid value's node label in schema, based on display name (i.e. shown to the user in a dropdown to fill in)
            req_val = req_val["userEnteredValue"]
            req_val_node_label = self.dmge.get_node_label(req_val)
            if not req_val_node_label:
                # if this node is not in the graph
                # continue - there are no dependencies for it
                continue
            # check if this required/valid value has additional dependency attributes
            val_dependencies = self.dmge.get_node_dependencies(
                req_val_node_label, schema_ordered=False
            )

            # prepare request calls
            dependency_formatting_body = {"requests": []}

            # set conditiaon formatting for dependencies.
            if val_dependencies:
                dependency_formatting_body["requests"] = self._dependency_formatting(
                    i,
                    req_val,
                    ordered_metadata_fields,
                    val_dependencies,
                    dependency_formatting_body,
                )

            if dependency_formatting_body["requests"]:
                requests_body["requests"].append(dependency_formatting_body["requests"])
        return requests_body

    def _create_requests_body(
        self,
        required_metadata_fields,
        ordered_metadata_fields,
        json_schema,
        spreadsheet_id,
        sheet_url,
        strict: Optional[bool],
    ):
        """Create and store all formatting changes for the google sheet to
        execute at once.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
            ordered_metadata_fields: List[list], contining all the attributes to
                add as columns, ordered
            json_schema: dict, representing a handful of values
                representing the data model, including: '$schema', '$id', 'title',
                'type', 'properties', 'required'
            spreadsheet_id: str, of the id for the google sheet
            sheet_url (Will be deprecated): a boolean ; determine if a pandas dataframe or a google sheet url gets return
            strict (Optional Bool): strictness with which to apply validation rules to google sheets. True, blocks incorrect entries, False, raises a warning
        Return:
            requests_body(dict):
                containing all the update requests to add to the gs
        """
        # store all requests to execute at once
        requests_body = {}
        requests_body["requests"] = []
        for i, req in enumerate(ordered_metadata_fields[0]):
            # Gather validation rules and valid values for attribute.
            validation_rules = self.dmge.get_node_validation_rules(
                node_display_name=req
            )
            if isinstance(validation_rules, dict):
                validation_rules = extract_component_validation_rules(
                    validation_rules_dict=validation_rules, manifest_component=self.root
                )

            # Add regex match validaiton rule to Google Sheets.
            if validation_rules and sheet_url:
                requests_body = self._request_regex_match_vr_formatting(
                    validation_rules, i, spreadsheet_id, requests_body, strict
                )

            if req in json_schema["properties"].keys():
                valid_values = self._get_valid_values_from_jsonschema_property(
                    json_schema["properties"][req]
                )
            else:
                valid_values = []

            # Set row formatting
            get_row_formatting = self._request_row_format(i, req)
            if get_row_formatting:
                requests_body["requests"].append(get_row_formatting)

            # set color of required columns to blue
            required_columns_color = self._set_required_columns_color(
                i, req, json_schema
            )
            if required_columns_color:
                requests_body["requests"].append(required_columns_color)
            # Add note on how to use multi-select, when appropriate
            note_vv = self._request_note_valid_values(
                i, req, validation_rules, valid_values
            )
            if note_vv:
                requests_body["requests"].append(note_vv)

            # Adding value-constraints, if any
            values = required_metadata_fields[req]
            req_vals = [{"userEnteredValue": value} for value in values if value]
            if not req_vals:
                continue

            # for attributes that don't require "list", create dropdown options and set up data validation rules
            if not rule_in_rule_list("list", validation_rules):
                create_dropdown = self._request_dropdown(
                    i, req_vals, spreadsheet_id, validation_rules, valid_values
                )
                if create_dropdown:
                    requests_body["requests"].append(create_dropdown)

            # for attributes that require "list", simply store valid values (if any) in second sheet
            elif len(req_vals) > 0 and rule_in_rule_list("list", validation_rules):
                self._store_valid_values_as_data_dictionary(i, req_vals, spreadsheet_id)

            # generate a conditional format rule for each required value (i.e. valid value)
            # for this field (i.e. if this field is set to a valid value that may require additional
            # fields to be filled in, these additional fields will be formatted in a custom style (e.g. red background)

            requests_body = self._request_dependency_formatting(
                i, req_vals, ordered_metadata_fields, requests_body
            )

        # Set borders formatting
        borders_formatting = self._request_cell_borders()
        if borders_formatting:
            requests_body["requests"].append(borders_formatting)
        return requests_body

    def _create_empty_gs(
        self,
        required_metadata_fields,
        json_schema,
        spreadsheet_id,
        sheet_url,
        strict: Optional[bool],
    ):
        """Generate requests to add columns and format the google sheet.
        Args:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
            json_schema: dict, representing a handful of values
                representing the data model, including: '$schema', '$id', 'title',
                'type', 'properties', 'required'
            spreadsheet_id: str, of the id for the google sheet
            sheet_url (str): google sheet url of template manifest
            strict (Optional Bool): strictness with which to apply validation rules to google sheets. True, blocks incorrect entries, False, raises a warning
        Returns:
            manifest_url (str): url of the google sheet manifest.
        """
        # adding columns to manifest sheet
        response, ordered_metadata_fields = self._gs_add_and_format_columns(
            required_metadata_fields, spreadsheet_id
        )

        # Add additional metadata
        response = self._gs_add_additional_metadata(
            required_metadata_fields, ordered_metadata_fields, spreadsheet_id
        )

        # Create requests body to add additional formatting to the sheets
        requests_body = self._create_requests_body(
            required_metadata_fields,
            ordered_metadata_fields,
            json_schema,
            spreadsheet_id,
            sheet_url,
            strict,
        )

        # Execute requests
        execute_google_api_requests(
            self.sheet_service,
            requests_body,
            service_type="batch_update",
            spreadsheet_id=spreadsheet_id,
        )

        # Setting up spreadsheet permissions (setup so that anyone with the link can edit)
        self._set_permissions(spreadsheet_id)

        # generating spreadsheet URL
        manifest_url = "https://docs.google.com/spreadsheets/d/" + spreadsheet_id
        return manifest_url

    def _gather_all_fields(self, fields, json_schema):
        """Gather all the attributes/fields to include as columns in the manifest.
        Args:
            fields(list[str]): fields/attributes to search
        Returns:
            required_metadata_fields(dict):
                keys: of all the fields/attributes that need to be added
                    to the manifest
                values(list[str]): valid values
        """
        # Get required fields
        required_metadata_fields = self._get_required_metadata_fields(
            json_schema, fields
        )
        # Add additional dependencies
        required_metadata_fields = self._gather_dependency_requirements(
            json_schema, required_metadata_fields
        )

        # Add additional metadata as entries to columns
        required_metadata_fields = self._get_additional_metadata(
            required_metadata_fields
        )
        return required_metadata_fields

    @tracer.start_as_current_span("ManifestGenerator::get_empty_manifest")
    def get_empty_manifest(
        self,
        strict: Optional[bool],
        json_schema_filepath: str = None,
        sheet_url: Optional[bool] = None,
    ):
        """Create an empty manifest using specifications from the
        json schema.
        Args:
            strict (bool): strictness with which to apply validation rules to google sheets. If true, blocks incorrect entries; if false, raises a warning
            json_schema_filepath (str): path to json schema file
            sheet_url (Will be deprecated): a boolean ; determine if a pandas dataframe or a google sheet url gets return
            strict (Optional Bool): strictness with which to apply validation rules to google sheets. True, blocks incorrect entries, False, raises a warning
        Returns:
            manifest_url (str): url of the google sheet manifest.
        TODO:
            Refactor to not be dependent on GS.
        """
        spreadsheet_id = self._create_empty_manifest_spreadsheet(self.title)
        json_schema = self._get_json_schema(json_schema_filepath=json_schema_filepath)

        required_metadata_fields = self._gather_all_fields(
            json_schema["properties"].keys(), json_schema
        )

        manifest_url = self._create_empty_gs(
            required_metadata_fields,
            json_schema,
            spreadsheet_id,
            sheet_url=sheet_url,
            strict=strict,
        )
        return manifest_url

    def _get_missing_columns(self, headers_1: list, headers_2: list) -> list:
        """Compare two colunm sets and get cols that are in headers_1, but not headers_2
        Args:
            headers_1 (list): list of column headers
            headers_2 (list): list of column headers
        Returns:
            list: column headers in headers_1 but not headers_2

        """
        return set(headers_1) - set(headers_2)

    @tracer.start_as_current_span("ManifestGenerator::set_dataframe_by_url")
    def set_dataframe_by_url(
        self,
        manifest_url: str,
        manifest_df: pd.DataFrame,
        out_of_schema_columns: set = None,
    ) -> ps.Spreadsheet:
        """Update Google Sheets using given pandas DataFrame.
        Args:
            manifest_url (str): Google Sheets URL.
            manifest_df (pd.DataFrame): Data frame to "upload".
            out_of_schema_columns (set): Columns that are in downloaded manifest, but not in current schema.
        Returns:
            ps.Spreadsheet: A Google Sheet object.
        """

        # authorize pygsheets to read from the given URL
        gc = ps.authorize(custom_credentials=self.creds)

        # open google sheets and extract first sheet
        # This sheet already contains headers.
        sh = gc.open_by_url(manifest_url)
        wb = sh[0]

        wb.set_dataframe(manifest_df, (1, 1), fit=True)

        # update validation rules (i.e. no validation rules) for out of schema columns, if any
        # TODO: similarly clear formatting for out of schema columns, if any
        if out_of_schema_columns:
            num_out_of_schema_columns = len(out_of_schema_columns)
            start_col = self._column_to_letter(
                len(manifest_df.columns) - num_out_of_schema_columns
            )  # find start of out of schema columns
            end_col = self._column_to_letter(
                len(manifest_df.columns) + 1
            )  # find end of out of schema columns
            wb.set_data_validation(start=start_col, end=end_col, condition_type=None)

        # set permissions so that anyone with the link can edit
        sh.share("", role="writer", type="anyone")

        return sh

    def get_dataframe_by_url(self, manifest_url: str) -> pd.DataFrame:
        """Retrieve pandas DataFrame from table in Google Sheets.

        Args:
            manifest_url (str): Google Sheets URL.

        Return:
            pd.DataFrame: Data frame corresponding to table in given URL.
        """

        # authorize pygsheets to read from the given URL
        gc = ps.authorize(custom_credentials=self.creds)

        # open google sheets and extract first sheet
        sh = gc.open_by_url(manifest_url)
        wb = sh[0]

        # get column headers and read it into a dataframe
        manifest_df = wb.get_as_df(hasHeader=True)

        # An empty column is sometimes included
        if "" in manifest_df:
            manifest_df.drop(columns=[""], inplace=True)

        return manifest_df

    def map_annotation_names_to_display_names(
        self, annotations: pd.DataFrame
    ) -> pd.DataFrame:
        """Update columns names to use display names for consistency.

        Args:
            annotations (pd.DataFrame): Annotations table.

        Returns:
            pd.DataFrame: Annotations table with updated column headers.
        """
        # Get list of attribute nodes from data model
        model_nodes = self.graph.nodes

        # Subset annotations to those appearing as a label in the model
        labels = filter(lambda x: x in model_nodes, annotations.columns)

        # Generate a dictionary mapping labels to display names
        label_map = {l: model_nodes[l]["displayName"] for l in labels}

        # Use the above dictionary to rename columns in question
        return annotations.rename(columns=label_map)

    @tracer.start_as_current_span("ManifestGenerator::get_manifest_with_annotations")
    def get_manifest_with_annotations(
        self, annotations: pd.DataFrame, strict: Optional[bool] = None
    ) -> Tuple[ps.Spreadsheet, pd.DataFrame]:
        """Generate manifest, optionally with annotations (if requested).
        Args:
            annotations (pd.DataFrame): Annotations table (can be empty).
            strict (Optional Bool): strictness with which to apply validation rules to google sheets. True, blocks incorrect entries, False, raises a warning
        Returns:
            Tuple[ps.Spreadsheet, pd.DataFrame]: Both the Google Sheet
            URL and the corresponding data frame is returned.
        """
        # Map annotation labels to display names to match manifest columns
        annotations = self.map_annotation_names_to_display_names(annotations)

        # Convert annotations table into dictionary, but maintain order
        annotations_dict_raw = annotations.to_dict(into=OrderedDict)
        annotations_dict = OrderedDict(
            (k, list(v.values())) for k, v in annotations_dict_raw.items()
        )

        # Needs to happen before get_empty_manifest() gets called
        self.additional_metadata = annotations_dict

        # Generate empty manifest using `additional_metadata`
        # With annotations added, regenerate empty manifest
        manifest_url = self.get_empty_manifest(sheet_url=True, strict=strict)
        manifest_df = self.get_dataframe_by_url(manifest_url=manifest_url)

        # Annotations clashing with manifest attributes are skipped
        # during empty manifest generation. For more info, search
        # for `additional_metadata` in `self.get_empty_manifest`.
        # Hence, the shared columns need to be updated separately.
        # This approach assumes that `update_df` returns
        # a data frame whose columns are in the same order
        manifest_df = update_df(manifest_df, annotations)
        manifest_sh = self.set_dataframe_by_url(manifest_url, manifest_df)
        manifest_url = manifest_sh.url

        return manifest_url, manifest_df

    @tracer.start_as_current_span("ManifestGenerator::export_sheet_to_excel")
    def export_sheet_to_excel(
        self, title: str = None, manifest_url: str = None, output_location: str = None
    ) -> str:
        """
        export manifest as an Excel spreadsheet and return local file path
        Args:
            title: title of the exported excel spreadsheet
            manifest_url: manifest google sheet url
            output_location: the location where the exported excel file would live
        return:
            Export manifest to a desired location.
        """
        # construct file name
        file_name = title + ".xlsx"
        # if file path exists and it contains a file name:
        if output_location:
            if os.path.exists(output_location):
                if (
                    Path(output_location).suffix == ".xlsx"
                    or Path(output_location).suffix == ".xls"
                ):
                    output_excel_file_path = output_location
                # if users define the location but it doesn't contain a file name, we should add the file name:
                else:
                    output_excel_file_path = os.path.join(output_location, file_name)

        # trigger a warning if file path is provided but does not exist
        elif output_location and not os.path.exists(output_location):
            output_excel_file_path = os.path.abspath(
                os.path.join(os.getcwd(), file_name)
            )
            logger.warning(
                f"{output_location} does not exist. Using current working directory {output_excel_file_path}"
            )
        # otherwise, use the default location
        else:
            output_excel_file_path = os.path.abspath(
                os.path.join(os.getcwd(), file_name)
            )

        # export the manifest to excel
        export_manifest_drive_service(
            manifest_url,
            file_path=output_excel_file_path,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        return output_excel_file_path

    @tracer.start_as_current_span("ManifestGenerator::_handle_output_format_logic")
    def _handle_output_format_logic(
        self,
        output_format: str = None,
        output_path: str = None,
        sheet_url: bool = None,
        empty_manifest_url: str = None,
        dataframe: pd.DataFrame = None,
        out_of_schema_columns: set = None,
    ):
        """
        Handle the logic between sheet_url parameter and output_format parameter to determine the type of output to return
        Args:
            output_format: Determines if Google sheet URL, pandas dataframe, or Excel spreadsheet gets returned.
            sheet_url (Will be deprecated): a boolean ; determine if a pandas dataframe or a google sheet url gets return
            empty_manifest_url: Google sheet URL that leads to an empty manifest
            dataframe: the pandas dataframe that contains the metadata that needs to be populated to an empty manifest
            output_path: Determines the output path of the exported manifest (only relevant if returning an excel spreadsheet)
            out_of_schema_columns (set): Columns that are in downloaded manifest, but not in current schema.
        Return:
            a pandas dataframe, file path of an excel spreadsheet, or a google sheet URL
        TODO:
            Depreciate sheet URL and add google_sheet as an output_format choice.

        ```mermaid
        flowchart TD
            A[Start] --> B{Output Format is 'dataframe'?}
            B -- Yes --> C[Return DataFrame]
            B -- No --> D{Output Format is 'excel'?}
            D -- Yes --> E[Export to Excel]
            E --> F[Populate Excel]
            F --> G[Return Excel Path]
            D -- No --> H{Sheet URL is set?}
            H -- Yes --> I[Set DataFrame by URL]
            I --> J[Return Sheet URL]
            H -- No --> K[Default Return DataFrame]
            C --> L[End]
            G --> L
            J --> L
            K --> L
        ```
        """

        # if the output type gets set to "dataframe", return a data frame
        if output_format == "dataframe":
            return dataframe

        # if the output type gets set to "excel", return an excel spreadsheet
        elif output_format == "excel":
            # export manifest url that only contains column headers to Excel
            output_file_path = self.export_sheet_to_excel(
                title=self.title,
                manifest_url=empty_manifest_url,
                output_location=output_path,
            )

            # populate an excel spreadsheet with the existing dataframe
            self.populate_existing_excel_spreadsheet(output_file_path, dataframe)

            return output_file_path

        # Return google sheet if sheet_url flag is raised.
        elif sheet_url:
            manifest_sh = self.set_dataframe_by_url(
                manifest_url=empty_manifest_url,
                manifest_df=dataframe,
                out_of_schema_columns=out_of_schema_columns,
            )
            return manifest_sh.url

        # Default return a DataFrame
        else:
            return dataframe

    @staticmethod
    @tracer.start_as_current_span("ManifestGenerator::create_single_manifest")
    def create_single_manifest(
        path_to_data_model: str,
        graph_data_model: nx.MultiDiGraph,
        data_type: str,
        access_token: Optional[str] = None,
        dataset_id: Optional[str] = None,
        strict: Optional[bool] = True,
        title: Optional[str] = None,
        output_format: Literal["google_sheet", "excel", "dataframe"] = "google_sheet",
        use_annotations: Optional[bool] = False,
    ) -> Union[str, pd.DataFrame]:
        """Create a single manifest

        Args:
            path_to_data_model (str): data model schema
            graph_data_model (nx.MultiDiGraph): graph data model
            data_type (str): data type of a manifest
            access_token (Optional[str], optional): synapse access token. Required when getting an existing manifest. Defaults to None.
            dataset_id (Optional[str], optional):dataset id when generating an existing manifest. Defaults to None. Defaults to None.
            strict (Optional[bool], optional): strictness with which to apply validation rules to google sheets. Defaults to True.
            title (Optional[str], optional):title of a given manifest. Defaults to None.
            output_format (Literal['google_sheet', 'excel', 'dataframe'], optional): format of manifest. Defaults to "google_sheet".
            use_annotations (Optional[bool], optional):whether to use annotations. Defaults to False.

        Returns:
            Union[str, pd.DataFrame]: Googlesheet URL or pandas dataframe or an excel file path
        """
        # create object of type ManifestGenerator
        manifest_generator = ManifestGenerator(
            path_to_data_model=path_to_data_model,
            graph=graph_data_model,
            title=title,
            root=data_type,
            use_annotations=use_annotations,
            alphabetize_valid_values="ascending",
        )

        # if returning a dataframe
        if output_format:
            if "dataframe" in output_format:
                output_format = "dataframe"

        result = manifest_generator.get_manifest(
            dataset_id=dataset_id,
            sheet_url=True,
            output_format=output_format,
            access_token=access_token,
            strict=strict,
        )

        return result

    @staticmethod
    @tracer.start_as_current_span("ManifestGenerator::create_manifests")
    def create_manifests(
        path_to_data_model: str,
        data_types: list,
        data_model_labels: DisplayLabelType = "class_label",
        access_token: Optional[str] = None,
        dataset_ids: Optional[list] = None,
        output_format: Literal["google_sheet", "excel", "dataframe"] = "google_sheet",
        title: Optional[str] = None,
        strict: Optional[bool] = True,
        use_annotations: Optional[bool] = False,
    ) -> Union[List[str], List[pd.DataFrame]]:
        """Create multiple manifests

        Args:
            path_to_data_model (str): str path to data model
            data_types (list): a list of data types
            access_token (str, optional): synapse access token. Required when getting an existing manifest. Defaults to None.
            dataset_ids (list, optional): a list of dataset ids when generating an existing manifest. Defaults to None.
            output_format (str, optional):format of manifest. It has three options: google sheet, excel or dataframe. Defaults to None.
            title (str, optional): title of a given manifest. Defaults to None.
            strict (bool, optional): strictness with which to apply validation rules to google sheets. Defaults to None.
            use_annotations (bool, optional): whether to use annotations. Defaults to False.

        Returns:
            Union[List[str], List[pd.DataFrame]]: a list of Googlesheet URLs, a list of pandas dataframes or excel file paths

        ```mermaid
        sequenceDiagram
            participant User
            participant Function
            participant DataModelParser
            participant DataModelGraph
            participant ManifestGenerator
            User->>Function: call create_manifests
            Function->>Function: check dataset_ids and validate inputs
            Function->>DataModelParser: parse data model
            DataModelParser-->>Function: return parsed data model
            Function->>DataModelGraph: generate graph
            DataModelGraph-->>Function: return graph data model
            alt data_types == "all manifests"
                loop for each component
                    Function->>ManifestGenerator: create manifest for component
                    ManifestGenerator-->>Function: single manifest
                end
            else
                loop for each data_type
                    Function->>ManifestGenerator: create single manifest
                    ManifestGenerator-->>Function: single manifest
                end
            end
            Function-->>User: return manifests based on output_format
        ```
        """
        if dataset_ids:
            # Check that the number of submitted data_types matches
            # the number of dataset_ids (if applicable)
            len_data_types = len(data_types)
            len_dataset_ids = len(dataset_ids)

            if len_data_types != len_dataset_ids:
                raise ValueError(
                    f"There is a mismatch in the number of data_types and dataset_id's that "
                    f"submitted. Please check your submission and try again."
                )

            # Raise an error if used in conjunction with datatype = 'all_manifests'
            if data_types[0] == "all manifests":
                raise ValueError(
                    "When submitting 'all manifests' as the data_type cannot also submit dataset_id. "
                    "Please check your submission and try again."
                )

        data_model_parser = DataModelParser(path_to_data_model=path_to_data_model)

        # Parse Model
        parsed_data_model = data_model_parser.parse_model()

        # Instantiate DataModelGraph
        data_model_grapher = DataModelGraph(parsed_data_model, data_model_labels)

        # Generate graph
        graph_data_model = data_model_grapher.graph

        # Gather all returned result urls
        all_results = []
        if data_types[0] == "all manifests":
            dmge = DataModelGraphExplorer(graph_data_model)
            component_digraph = dmge.get_digraph_by_edge_type("requiresComponent")
            components = component_digraph.nodes()
            for component in components:
                if title:
                    t = f"{title}.{component}.manifest"
                else:
                    t = f"Example.{component}.manifest"
                if output_format != "excel":
                    result = ManifestGenerator.create_single_manifest(
                        path_to_data_model=path_to_data_model,
                        data_type=component,
                        graph_data_model=graph_data_model,
                        output_format=output_format,
                        title=t,
                        strict=strict,
                        access_token=access_token,
                    )
                    all_results.append(result)
                else:
                    logger.error(
                        "Currently we do not support returning multiple files as Excel format at once. Please choose a different output format. "
                    )
        else:
            for i, dt in enumerate(data_types):
                if not title:
                    t = f"Example.{dt}.manifest"
                else:
                    if len(data_types) > 1:
                        t = f"{title}.{dt}.manifest"
                    else:
                        t = title
                if dataset_ids:
                    # if a dataset_id is provided add this to the function call.
                    result = ManifestGenerator.create_single_manifest(
                        path_to_data_model=path_to_data_model,
                        data_type=dt,
                        graph_data_model=graph_data_model,
                        dataset_id=dataset_ids[i],
                        output_format=output_format,
                        title=t,
                        strict=strict,
                        access_token=access_token,
                        use_annotations=use_annotations,
                    )
                else:
                    result = ManifestGenerator.create_single_manifest(
                        path_to_data_model=path_to_data_model,
                        data_type=dt,
                        graph_data_model=graph_data_model,
                        output_format=output_format,
                        title=t,
                        strict=strict,
                        access_token=access_token,
                        use_annotations=use_annotations,
                    )

                # if output is pandas dataframe or google sheet url
                if isinstance(result, str) or isinstance(result, pd.DataFrame):
                    all_results.append(result)
                else:
                    if len(data_types) > 1:
                        logger.warning(
                            f"Currently we do not support returning multiple files as Excel format at once. Only {t} would get returned. "
                        )
                    return result

        return all_results

    @tracer.start_as_current_span("ManifestGenerator::get_manifest")
    def get_manifest(
        self,
        dataset_id: str = None,
        sheet_url: bool = None,
        json_schema: str = None,
        output_format: str = None,
        output_path: str = None,
        access_token: str = None,
        strict: Optional[bool] = None,
    ) -> Union[str, pd.DataFrame]:
        """Gets manifest for a given dataset on Synapse.
           TODO: move this function to class MetadatModel (after MetadataModel is refactored)

        Args:
            dataset_id: Synapse ID of the "dataset" entity on Synapse (for a given center/project).
            sheet_url (Will be deprecated): a boolean ; determine if a pandas dataframe or a google sheet url gets return
            output_format: Determines if Google sheet URL, pandas dataframe, or Excel spreadsheet gets returned.
            output_path: Determines the output path of the exported manifest
            access_token: Token in .synapseConfig. Since we could not pre-load access_token as an environment variable on AWS, we have to add this variable.

        Returns:
            Googlesheet URL, pandas dataframe, or an Excel spreadsheet

        ```mermaid
        flowchart TD
            Start[Start] --> DatasetIDCheck{Dataset ID provided?}
            DatasetIDCheck -- No --> EmptyManifestURL[Get Empty Manifest URL]
            EmptyManifestURL --> OutputFormatCheck{Output Format is 'excel'?}
            OutputFormatCheck -- Yes --> ExportToExcel[Export to Excel]
            OutputFormatCheck -- No --> ReturnManifestURL[Return Manifest URL]
            DatasetIDCheck -- Yes --> InstantiateSynapseStorage[Instantiate SynapseStorage]
            InstantiateSynapseStorage --> UpdateManifestFiles[Update Dataset Manifest Files]
            UpdateManifestFiles --> GetEmptyManifestURL[Get Empty Manifest URL]
            GetEmptyManifestURL --> ManifestRecordCheck{Manifest Record exists?}
            ManifestRecordCheck -- Yes --> UpdateDataframe[Update Dataframe]
            UpdateDataframe --> HandleOutputFormatLogic[Handle Output Format Logic]
            HandleOutputFormatLogic --> ReturnResult[Return Result]
            ManifestRecordCheck -- No --> UseAnnotationsCheck{Use Annotations?}

            UseAnnotationsCheck -- No --> CreateDataframe[Create dataframe from empty manifest on Google]
            CreateDataframe --> ManifestFileBasedCheck1{Manifest file-based?}
            ManifestFileBasedCheck1 -- Yes --> AddEntityID[Add entityId and filename to manifest df]
            ManifestFileBasedCheck1 -- No --> UseDataframe[Use dataframe from an empty manifest]

            AddEntityID --> HandleOutputFormatLogic
            UseDataframe --> HandleOutputFormatLogic

            UseAnnotationsCheck -- Yes --> ManifestFileBasedCheck2{Manifest file-based?}
            ManifestFileBasedCheck2 -- No --> HandleOutputFormatLogic
            ManifestFileBasedCheck2 -- Yes --> ProcessAnnotations[Process Annotations]
            ProcessAnnotations --> AnnotationsEmptyCheck{Annotations Empty?}
            AnnotationsEmptyCheck -- Yes --> CreateDataframeFromEmpty[Create dataframe from an empty manifest on Google]
            CreateDataframeFromEmpty --> UpdateDataframeWithAnnotations[Update dataframe]
            AnnotationsEmptyCheck -- No --> GetManifestWithAnnotations[Get Manifest with Annotations]
            GetManifestWithAnnotations --> UpdateDataframeWithAnnotations
            UpdateDataframeWithAnnotations --> HandleOutputFormatLogic
            ReturnResult --> End[End]
            ReturnManifestURL --> End
            ExportToExcel --> End
        ```
        """
        # Handle case when no dataset ID is provided
        if not dataset_id:
            manifest_url = self.get_empty_manifest(
                json_schema_filepath=json_schema, strict=strict, sheet_url=sheet_url
            )

            # if output_form parameter is set to "excel", return an excel spreadsheet
            if output_format == "excel":
                output_file_path = self.export_sheet_to_excel(
                    title=self.title,
                    manifest_url=manifest_url,
                    output_location=output_path,
                )
                return output_file_path
            # since we are not going to return an empty dataframe for an empty manifest, here we will just return a google sheet url for all other cases
            else:
                return manifest_url

        # Otherwise, create manifest using the given dataset
        # TODO: avoid explicitly exposing Synapse store functionality
        # just instantiate a Store class and let it decide at runtime/config
        # the store type
        if access_token:
            # for getting an existing manifest on AWS
            store = SynapseStorage(access_token=access_token)
        else:
            store = SynapseStorage()

        # Get manifest file associated with given dataset (if applicable)
        # populate manifest with set of new files (if applicable)
        manifest_record = store.updateDatasetManifestFiles(
            self.dmge, datasetId=dataset_id, store=False
        )

        # get URL of an empty manifest file created based on schema component
        empty_manifest_url = self.get_empty_manifest(strict=strict, sheet_url=True)

        # Populate empty template with existing manifest
        if manifest_record:
            # TODO: Update or remove the warning in self.__init__() if
            # you change the behavior here based on self.use_annotations
            # Update df with existing manifest. Agnostic to output format
            updated_df, out_of_schema_columns = self._update_dataframe_with_existing_df(
                empty_manifest_url=empty_manifest_url, existing_df=manifest_record[1]
            )

            # determine the format of manifest
            result = self._handle_output_format_logic(
                output_format=output_format,
                output_path=output_path,
                sheet_url=sheet_url,
                empty_manifest_url=empty_manifest_url,
                dataframe=updated_df,
                out_of_schema_columns=out_of_schema_columns,
            )
            return result

        # Generate empty template and optionally fill in with annotations
        # if there is no existing manifest and use annotations is set to True,
        # pull annotations (in reality, annotations should be empty when there is no existing manifest)
        else:
            # Using getDatasetAnnotations() to retrieve file names and subset
            # entities to files and folders (ignoring tables/views)
            annotations = pd.DataFrame()
            if self.use_annotations:
                if self.is_file_based:
                    annotations = store.getDatasetAnnotations(dataset_id)
                    # Update `additional_metadata` and generate manifest
                    manifest_url, manifest_df = self.get_manifest_with_annotations(
                        annotations, strict=strict
                    )

                # If the annotations are empty,
                # ie if there are no annotations to pull or annotations were unable to be pulled because the metadata is not file based,
                # then create manifest from an empty manifest
                if annotations.empty:
                    empty_manifest_df = self.get_dataframe_by_url(empty_manifest_url)
                    manifest_df = empty_manifest_df

                    logger.warning(
                        f"Annotations were not able to be gathered for the given parameters. This manifest will be generated from an empty manifest."
                    )

            else:
                empty_manifest_df = self.get_dataframe_by_url(empty_manifest_url)
                if self.is_file_based:
                    # for file-based manifest, make sure that entityId column and Filename column still gets filled even though use_annotations gets set to False
                    manifest_df = store.add_entity_id_and_filename(
                        dataset_id, empty_manifest_df
                    )
                else:
                    manifest_df = empty_manifest_df

            # Update df with existing manifest. Agnostic to output format
            updated_df, out_of_schema_columns = self._update_dataframe_with_existing_df(
                empty_manifest_url=empty_manifest_url, existing_df=manifest_df
            )

            # determine the format of manifest that gets return
            result = self._handle_output_format_logic(
                output_format=output_format,
                output_path=output_path,
                sheet_url=sheet_url,
                empty_manifest_url=empty_manifest_url,
                dataframe=updated_df,
                out_of_schema_columns=out_of_schema_columns,
            )
            return result

    def _get_end_columns(
        self, current_schema_headers, existing_manifest_headers, out_of_schema_columns
    ):
        """
        Gather columns to be added to the end of the manifest, and ensure entityId is at the end.
        Args:
            current_schema_headers: list, columns in the current manifest schema
            existing_manifest_headers: list, columns in the existing manifest
            out_of_schema_columns: set, columns that are in the existing manifest, but not the current schema
        Returns:
            end_columns: list of columns to be added to the end of the manifest.
        """
        # Identify columns to add to the end of the manifest
        end_columns = list(out_of_schema_columns)

        # Make sure want Ids are placed at end of manifest, in given order.
        for id_name in ["Uuid", "Id", "entityId"]:
            if id_name in end_columns:
                end_columns.remove(id_name)
                end_columns.append(id_name)

        # Add entity_id to the end columns if it should be there but isn't
        if (
            "entityId" in (current_schema_headers or existing_manifest_headers)
            and "entityId" not in end_columns
        ):
            end_columns.append("entityId")
        return end_columns

    def _update_dataframe_with_existing_df(
        self, empty_manifest_url: str, existing_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Handle scenario when existing manifest does not match new manifest template due to changes in the data model:
            the sheet column header reflect the latest schema the existing manifest column-set may be outdated
            ensure that, if missing, attributes from the latest schema are added to the column-set of the existing manifest so that the user can modify their data if needed
            to comply with the latest schema.
        Args:
            empty_manifest_url (str): Google Sheet URL with an empty manifest with headers.
            existing_df (Pd.DataFrame): df of an existing manifest

        Returns:
            updated_df (Pd.DataFrame): existing_df with new_columns added.
            out_of_schema_columns (set): Columns that are in downloaded manifest, but not in current schema.
        """

        # Get headers for the current schema and existing manifest df.
        current_schema_headers = list(
            self.get_dataframe_by_url(manifest_url=empty_manifest_url).columns
        )
        existing_manifest_headers = list(existing_df.columns)

        # Find columns that exist in the current schema, but are not in the manifest being downloaded.
        new_columns = self._get_missing_columns(
            current_schema_headers, existing_manifest_headers
        )

        # Find columns that exist in the manifest being downloaded, but not in the current schema.
        out_of_schema_columns = self._get_missing_columns(
            existing_manifest_headers, current_schema_headers
        )

        # clean empty columns if any are present (there should be none)
        # TODO: Remove this line once we start preventing empty column names
        if "" in new_columns:
            new_columns = new_columns.remove("")

        # Copy the df for updating.
        updated_df = existing_df.copy(deep=True)

        # update existing manifest w/ missing columns, if any
        if new_columns:
            updated_df = updated_df.assign(
                **dict(zip(new_columns, len(new_columns) * [""]))
            )

        end_columns = self._get_end_columns(
            current_schema_headers=current_schema_headers,
            existing_manifest_headers=existing_manifest_headers,
            out_of_schema_columns=out_of_schema_columns,
        )

        # sort columns in the updated manifest:
        # match latest schema order
        # move obsolete columns at the end
        updated_df = updated_df[self.sort_manifest_fields(updated_df.columns)]

        # move obsolete columns at the end with entityId at the very end
        updated_df = updated_df[
            [c for c in updated_df if c not in end_columns] + list(end_columns)
        ]
        return updated_df, out_of_schema_columns

    def _format_new_excel_column(self, worksheet, new_column_index: int, col: str):
        """Add new column to an openpyxl worksheet and format header.
        Args:
            worksheet: openpyxl worksheet
            new_column_index, int: index to add new column
        Return:
            modified worksheet
        """
        # Add column header
        worksheet.cell(row=1, column=new_column_index + 1).value = col
        # Format new column header
        worksheet.cell(row=1, column=new_column_index + 1).font = Font(
            size=8, bold=True, color="FF000000"
        )
        worksheet.cell(row=1, column=new_column_index + 1).alignment = Alignment(
            horizontal="center", vertical="bottom"
        )
        worksheet.cell(row=1, column=new_column_index + 1).fill = PatternFill(
            start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid"
        )
        return worksheet

    @tracer.start_as_current_span(
        "ManifestGenerator::populate_existing_excel_spreadsheet"
    )
    def populate_existing_excel_spreadsheet(
        self, existing_excel_path: str = None, additional_df: pd.DataFrame = None
    ):
        """Populate an existing excel spreadsheet by using an additional dataframe (to avoid sending metadata directly to Google APIs)
        New columns will be placed at the end of the spreadsheet.
        Args:
            existing_excel_path: path of an existing excel spreadsheet
            additional_df: additional dataframe
        Return:
            added new dataframe to the existing excel path.
        Note:
            - Done by rows and column as a way to preserve formatting.
            Doing a complete replacement will remove all conditional formatting and dropdowns.
        """
        # load workbook
        workbook = load_workbook(existing_excel_path)
        worksheet = workbook.active

        # Add new data to existing excel
        if not additional_df.empty:
            existing_excel_headers = [
                cell.value for cell in worksheet[1] if cell.value != None
            ]

            new_column_index = len(existing_excel_headers)
            df_columns = additional_df.columns

            # Iteratively fill workbook with contents of additional_df
            for row_num, row_contents in enumerate(
                dataframe_to_rows(additional_df, index=False, header=False), 2
            ):
                for index, col in enumerate(df_columns):
                    if col in existing_excel_headers:
                        # Get index of column header in existing excel to ensure no values are placed in incorrect spot.
                        existing_column_index = existing_excel_headers.index(col)
                        worksheet.cell(
                            row=row_num, column=existing_column_index + 1
                        ).value = row_contents[index]
                    else:
                        # Add new col to excel worksheet and format.
                        worksheet = self._format_new_excel_column(
                            worksheet=worksheet,
                            new_column_index=new_column_index,
                            col=col,
                        )
                        # Add data to column
                        worksheet.cell(
                            row=row_num, column=new_column_index + 1
                        ).value = row_contents[index]
                        # Add new column to headers so it can be accounted for.
                        existing_excel_headers.append(col)
                        # Update index for adding new columns.
                        new_column_index += 1
        workbook.save(existing_excel_path)

    def populate_manifest_spreadsheet(
        self,
        existing_manifest_path: str = None,
        empty_manifest_url: str = None,
        return_excel: bool = False,
        title: str = None,
    ):
        """Creates a google sheet manifest based on existing manifest.
        Args:
            existing_manifest_path: the location of the manifest containing metadata presently stored
            empty_manifest_url: the path to a manifest template to be prepopulated with existing's manifest metadata
            return_excel: if true, return an Excel spreadsheet instead of Google sheet
            title: title of output manifest
        """

        # read existing manifest
        manifest = load_df(existing_manifest_path)

        if return_excel:
            """if we are returning an Excel spreadsheet, do not populate dataframe to google"""
            # get an empty manifest
            manifest_url = empty_manifest_url

            # export the manifest to excel
            output_excel_file_path = self.export_sheet_to_excel(
                manifest_url=manifest_url, title=title
            )

            # populate exported sheet
            self.populate_existing_excel_spreadsheet(output_excel_file_path, manifest)
            return output_excel_file_path
        else:
            manifest_sh = self.set_dataframe_by_url(empty_manifest_url, manifest)
            return manifest_sh.url

    def sort_manifest_fields(self, manifest_fields, order="schema"):
        # order manifest fields alphabetically (base order)
        manifest_fields = sorted(manifest_fields)

        if order == "alphabetical":
            # if the order is alphabetical ensure that filename is first, if present
            if "Filename" in manifest_fields:
                manifest_fields.remove("Filename")
                manifest_fields.insert(0, "Filename")

        # order manifest fields based on data-model schema
        if order == "schema":
            if self.dmge and self.root:
                # get display names of dependencies
                dependencies_display_names = self.dmge.get_node_dependencies(self.root)

                # reorder manifest fields so that root dependencies are first and follow schema order
                manifest_fields = sorted(
                    manifest_fields,
                    key=lambda x: dependencies_display_names.index(x)
                    if x in dependencies_display_names
                    else len(manifest_fields) - 1,
                )
            else:
                raise ValueError(
                    f"Provide valid data model path and valid component from data model."
                )

        # always have entityId as last columnn, if present
        if "entityId" in manifest_fields:
            manifest_fields.remove("entityId")
            manifest_fields.append("entityId")

        return manifest_fields
